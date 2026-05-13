import logging
import re
import unicodedata
from io import BytesIO
from datetime import date, datetime
from pathlib import Path

from django.db.models import Q
from django.utils import timezone
from openpyxl import load_workbook

from .models import TicketAttendance, TicketAutoPauseReview
from users.access import is_ti_user

logger = logging.getLogger(__name__)

MONTH_TOKENS = {
    1: ('jan', 'janeiro', '01'),
    2: ('fev', 'fevereiro', '02'),
    3: ('mar', 'marco', '03'),
    4: ('abr', 'abril', '04'),
    5: ('mai', 'maio', '05'),
    6: ('jun', 'junho', '06'),
    7: ('jul', 'julho', '07'),
    8: ('ago', 'agosto', '08'),
    9: ('set', 'setembro', '09'),
    10: ('out', 'outubro', '10'),
    11: ('nov', 'novembro', '11'),
    12: ('dez', 'dezembro', '12'),
}


def _normalize(value: str) -> str:
    raw = (value or '').strip().lower()
    base = unicodedata.normalize('NFKD', raw).encode('ascii', 'ignore').decode('ascii')
    base = re.sub(r'\s+', ' ', base)
    return base


def _build_header_map(cells: list[str]) -> dict[str, int]:
    wanted = {
        'ti': None,
        'data': None,
        'contato': None,
        'setor': None,
        'notificacao': None,
        'prioridade': None,
        'falha': None,
        'acao': None,
        'fechado': None,
        'tempo': None,
        'acao_eficaz': None,
    }
    for idx, raw in enumerate(cells, start=1):
        key = _normalize(str(raw or ''))
        if key == 'ti':
            wanted['ti'] = idx
        elif key == 'data':
            wanted['data'] = idx
        elif key == 'contato':
            wanted['contato'] = idx
        elif key == 'setor':
            wanted['setor'] = idx
        elif key == 'notificacao':
            wanted['notificacao'] = idx
        elif key == 'prioridade':
            wanted['prioridade'] = idx
        elif key == 'falha':
            wanted['falha'] = idx
        elif key in {'acao / correcao', 'acao/correcao', 'acao correcao'}:
            wanted['acao'] = idx
        elif key == 'fechado':
            wanted['fechado'] = idx
        elif key == 'tempo':
            wanted['tempo'] = idx
        elif key == 'acao eficaz':
            wanted['acao_eficaz'] = idx
    return wanted


def _resolve_sheet(workbook, event_dt: datetime):
    best_sheet = workbook.active
    best_score = -1
    month_tokens = MONTH_TOKENS.get(event_dt.month, ())
    year_text = str(event_dt.year)
    for sheet in workbook.worksheets:
        normalized = _normalize(sheet.title)
        score = 0
        if year_text in normalized:
            score += 3
        if any(token in normalized for token in month_tokens):
            score += 3
        if score > best_score:
            best_score = score
            best_sheet = sheet
    return best_sheet


def _find_header(sheet):
    for row_idx in range(1, 8):
        raw = [sheet.cell(row=row_idx, column=col).value for col in range(1, 30)]
        header_map = _build_header_map(raw)
        if header_map['data'] and header_map['contato'] and header_map['notificacao']:
            return row_idx, header_map

    return 1, {
        'ti': 1,
        'data': 2,
        'contato': 3,
        'setor': 4,
        'notificacao': 5,
        'prioridade': 6,
        'falha': 7,
        'acao': 8,
        'fechado': 9,
        'tempo': 10,
        'acao_eficaz': 11,
    }


def _find_next_row(sheet, header_row: int, header_map: dict[str, int]) -> int:
    key_cols = [header_map[k] for k in ('data', 'contato', 'notificacao', 'fechado') if header_map.get(k)]
    if not key_cols:
        return header_row + 1
    row = header_row + 1
    while True:
        has_value = any((sheet.cell(row=row, column=col).value not in (None, '')) for col in key_cols)
        if not has_value:
            return row
        row += 1


def _format_dt(dt: datetime) -> str:
    return timezone.localtime(dt).strftime('%d/%m/%Y %H:%M')


def _format_workbook_dt(value) -> str:
    if value in (None, ''):
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d/%m/%Y %H:%M')
    return str(value).strip()


def _pending_auto_pause_reviews_count(attendant) -> int:
    return TicketAutoPauseReview.objects.filter(
        attendance__attendant=attendant,
        completed_at__isnull=True,
    ).count()


def _ticket_id_from_cell(value):
    if value in (None, ''):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)

    normalized = str(value).strip().lstrip('#')
    if normalized.isdigit():
        return int(normalized)
    return None


def _existing_attendance_keys_in_workbook(workbook) -> set[tuple[int, str]]:
    attendance_keys = set()
    for sheet in workbook.worksheets:
        header_row, header_map = _find_header(sheet)
        ticket_col = header_map.get('ti')
        closed_col = header_map.get('fechado')
        if not ticket_col:
            continue

        for row_idx in range(header_row + 1, sheet.max_row + 1):
            cell_ticket_id = _ticket_id_from_cell(sheet.cell(row=row_idx, column=ticket_col).value)
            if cell_ticket_id is not None:
                closed_value = _format_workbook_dt(sheet.cell(row=row_idx, column=closed_col).value) if closed_col else ''
                attendance_keys.add((cell_ticket_id, closed_value))
    return attendance_keys


def _month_range(export_month: date) -> tuple[datetime, datetime]:
    year = export_month.year
    month = export_month.month
    next_year = year + 1 if month == 12 else year
    next_month = 1 if month == 12 else month + 1
    current_tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime(year, month, 1, 0, 0), current_tz)
    end = timezone.make_aware(datetime(next_year, next_month, 1, 0, 0), current_tz)
    return start, end


def _eligible_attendances(attendant, export_month: date) -> list[TicketAttendance]:
    start, end = _month_range(export_month)
    return list(
        TicketAttendance.objects.filter(
            attendant=attendant,
            ended_at__isnull=False,
            ended_at__gte=start,
            ended_at__lt=end,
        )
        .filter(Q(auto_pause_review__isnull=True) | Q(auto_pause_review__completed_at__isnull=False))
        .select_related('ticket__created_by', 'attendant')
        .order_by('ticket_id', 'ended_at', 'id')
    )


def _format_duration_seconds(total_seconds: int) -> str:
    safe_seconds = max(int(total_seconds or 0), 0)
    minutes = safe_seconds // 60
    hours = minutes // 60
    mins = minutes % 60
    return f'{hours:02d}:{mins:02d}'


def _build_ticket_export_rows(attendant, existing_attendance_keys: set[tuple[int, str]], export_month: date) -> list[dict]:
    rows = []
    for attendance in _eligible_attendances(attendant, export_month):
        closed_at = _format_dt(attendance.ended_at)
        if (attendance.ticket_id, closed_at) in existing_attendance_keys:
            continue
        total_seconds = max(int((attendance.ended_at - attendance.started_at).total_seconds()), 0)
        rows.append(
            {
                'ticket': attendance.ticket,
                'attendant': attendant,
                'started_at': attendance.started_at,
                'ended_at': attendance.ended_at,
                'note': (attendance.note or '').strip(),
                'duration': _format_duration_seconds(total_seconds),
                'attendance_ids': [attendance.id],
            }
        )

    return sorted(rows, key=lambda item: (item['ended_at'], item['ticket'].id))


def _write_ticket_rows_to_workbook(workbook, rows: list[dict], export_month: date) -> None:
    sheet_reference_dt = timezone.make_aware(datetime(export_month.year, export_month.month, 1, 0, 0), timezone.get_current_timezone())
    for row in rows:
        sheet = _resolve_sheet(workbook, sheet_reference_dt)
        header_row, header_map = _find_header(sheet)
        target_row = _find_next_row(sheet, header_row, header_map)
        ticket = row['ticket']

        values = {
            'ti': ticket.id,
            'data': _format_dt(row['started_at']),
            'contato': _contact_name_for_ticket(ticket),
            'setor': _department_label_for_ticket(ticket),
            'notificacao': ticket.description or '',
            'prioridade': ticket.get_priority_display(),
            'falha': ticket.get_failure_type_display(),
            'acao': row['note'],
            'fechado': _format_dt(row['ended_at']),
            'tempo': row['duration'],
            'acao_eficaz': '',
        }

        for key, col in header_map.items():
            if not col or key not in values:
                continue
            sheet.cell(row=target_row, column=col, value=values[key])


def _contact_name_for_ticket(ticket) -> str:
    creator = ticket.created_by
    if not creator:
        return '-'
    full_name = creator.get_full_name().strip()
    return full_name or creator.username or '-'


def _department_label_for_ticket(ticket) -> str:
    creator = ticket.created_by
    if not creator:
        return ''
    if is_ti_user(creator):
        return 'TI'

    for attr_name in ('department', 'setor', 'ldap_department'):
        value = (getattr(creator, attr_name, '') or '').strip()
        if value:
            return value

    profile = getattr(creator, 'profile', None)
    if profile:
        for attr_name in ('department', 'setor', 'ldap_department'):
            value = (getattr(profile, attr_name, '') or '').strip()
            if value:
                return value

    return ''


def _mark_export_rows(rows: list[dict], exported_path: str) -> None:
    attendance_ids = [
        attendance_id
        for row in rows
        for attendance_id in row['attendance_ids']
    ]
    if not attendance_ids:
        return
    now = timezone.now()
    TicketAttendance.objects.filter(id__in=attendance_ids).update(
        exported_at=now,
        exported_path=exported_path,
    )


def _spreadsheet_export_blocker(attendant):
    pending_reviews = _pending_auto_pause_reviews_count(attendant)
    if pending_reviews:
        return (
            False,
            0,
            'Existem pausas automaticas pendentes para este atendente. '
            'Conclua essas revisoes antes de preencher a planilha.',
        )
    return None


def export_attendant_logs_to_uploaded_workbook(*, attendant, uploaded_file, export_month: date) -> tuple[bool, int, str, bytes | None, str]:
    blocker = _spreadsheet_export_blocker(attendant)
    if blocker:
        ok, count, detail = blocker
        return ok, count, detail, None, ''

    original_name = Path(getattr(uploaded_file, 'name', '') or 'chamados.xlsx').name
    if not original_name.lower().endswith('.xlsx'):
        return False, 0, 'Selecione uma planilha no formato .xlsx.', None, ''

    month_label = export_month.strftime('%m/%Y')
    try:
        wb = load_workbook(uploaded_file)
        rows = _build_ticket_export_rows(attendant, _existing_attendance_keys_in_workbook(wb), export_month)
        if not rows:
            return True, 0, f'Nenhum atendimento novo para exportar em {month_label}. Todos os atendimentos do atendente ja constam na planilha.', None, ''
        _write_ticket_rows_to_workbook(wb, rows, export_month)
        output = BytesIO()
        wb.save(output)
    except Exception as exc:
        logger.exception('Falha ao preencher planilha enviada para %s', attendant.username)
        return False, 0, f'Falha ao preencher planilha enviada: {exc}', None, ''

    download_name = f'preenchida-{original_name}'
    _mark_export_rows(rows, f'upload:{original_name}')
    return (
        True,
        len(rows),
        f'{len(rows)} atendimento(s) novo(s) de {month_label} exportado(s) com sucesso.',
        output.getvalue(),
        download_name,
    )
