import csv
import unicodedata
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from chamados.models import PhoneExtension


EXPECTED_COLUMNS = {
    'department': {'departamento', 'setor'},
    'name': {'colaborador', 'nome'},
    'phone': {'telefone', 'fone'},
    'extension': {'ramal'},
    'email': {'email', 'e-mail', 'mail'},
}


def _normalize(value):
    text = str(value or '').strip().lower()
    text = unicodedata.normalize('NFKD', text).encode('ascii', 'ignore').decode('ascii')
    return ' '.join(text.replace('_', ' ').replace('-', ' ').split())


def _clean(value):
    return str(value or '').strip()


class Command(BaseCommand):
    help = 'Importa ramais a partir de CSV ou XLSX com colunas Departamento, Colaborador, Telefone, Ramal e Email.'

    def add_arguments(self, parser):
        parser.add_argument('--source', required=True, help='Caminho do arquivo .xlsx ou .csv de ramais.')
        parser.add_argument('--created-by', default='fabiano.polone', help='Usuario que ficara como criador dos ramais.')
        parser.add_argument('--replace', action='store_true', help='Apaga os ramais atuais antes de importar.')
        parser.add_argument('--dry-run', action='store_true', help='Simula a importacao sem gravar no banco.')

    def handle(self, *args, **options):
        source = Path(options['source']).expanduser()
        if not source.exists():
            raise CommandError(f'Arquivo nao encontrado: {source}')

        created_by = self._resolve_user(options['created_by'])
        rows = self._read_rows(source)
        if not rows:
            raise CommandError('Nenhum ramal valido foi encontrado no arquivo.')

        created = 0
        updated = 0
        unchanged = 0

        self.stdout.write(f'Arquivo: {source}')
        self.stdout.write(f'Ramais validos encontrados: {len(rows)}')
        self.stdout.write(f'Criador: {created_by.username}')
        self.stdout.write(f'Modo: {"simulacao" if options["dry_run"] else "execucao"}')

        with transaction.atomic():
            if options['replace']:
                deleted_count, _ = PhoneExtension.objects.all().delete()
                self.stdout.write(f'Ramais atuais removidos: {deleted_count}')

            for row in rows:
                lookup = self._lookup_for_row(row)
                existing = PhoneExtension.objects.filter(**lookup).first()
                if existing is None:
                    PhoneExtension.objects.create(**row, created_by=created_by)
                    created += 1
                    continue

                changed_fields = []
                for field, value in row.items():
                    if getattr(existing, field) != value:
                        setattr(existing, field, value)
                        changed_fields.append(field)
                if changed_fields:
                    existing.created_by = existing.created_by or created_by
                    existing.save(update_fields=[*changed_fields, 'updated_at'])
                    updated += 1
                else:
                    unchanged += 1

            if options['dry_run']:
                transaction.set_rollback(True)

        self.stdout.write(self.style.SUCCESS('Importacao de ramais concluida.'))
        self.stdout.write(f'  criados: {created}')
        self.stdout.write(f'  atualizados: {updated}')
        self.stdout.write(f'  sem alteracao: {unchanged}')

    def _resolve_user(self, username):
        user_model = get_user_model()
        user = user_model.objects.filter(username__iexact=username).first()
        if user:
            return user
        user = user_model.objects.filter(is_superuser=True).order_by('id').first()
        if user:
            return user
        user = user_model.objects.order_by('id').first()
        if user:
            return user
        raise CommandError('Nao existe usuario no banco para vincular como criador dos ramais.')

    def _read_rows(self, source):
        suffix = source.suffix.lower()
        if suffix == '.xlsx':
            raw_rows = self._read_xlsx_rows(source)
        elif suffix == '.csv':
            raw_rows = self._read_csv_rows(source)
        else:
            raise CommandError('Formato nao suportado. Use .xlsx ou .csv.')

        header = next(raw_rows, None)
        if not header:
            return []
        column_map = self._build_column_map(header)
        missing = [field for field in ('department', 'name', 'phone', 'extension', 'email') if field not in column_map]
        if missing:
            expected = ', '.join(['Departamento', 'Colaborador', 'Telefone', 'Ramal', 'Email'])
            raise CommandError(f'Colunas obrigatorias ausentes. Esperado: {expected}.')

        rows = []
        for raw in raw_rows:
            row = {
                field: _clean(raw[index]) if index < len(raw) else ''
                for field, index in column_map.items()
            }
            if not any(row.values()):
                continue
            if not row['name'] and not row['extension'] and not row['email']:
                continue
            rows.append(row)
        return rows

    def _read_xlsx_rows(self, source):
        workbook = load_workbook(source, data_only=True)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only=True):
            yield list(row)

    def _read_csv_rows(self, source):
        raw = source.read_text(encoding='utf-8-sig')
        sample = raw[:2048]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=';,')
        except csv.Error:
            dialect = csv.excel
        yield from csv.reader(raw.splitlines(), dialect)

    def _build_column_map(self, header):
        column_map = {}
        for index, label in enumerate(header):
            normalized = _normalize(label)
            for field, aliases in EXPECTED_COLUMNS.items():
                if normalized in aliases:
                    column_map[field] = index
        return column_map

    def _lookup_for_row(self, row):
        if row['email']:
            return {'email__iexact': row['email']}
        return {
            'department__iexact': row['department'],
            'name__iexact': row['name'],
            'extension__iexact': row['extension'],
        }
