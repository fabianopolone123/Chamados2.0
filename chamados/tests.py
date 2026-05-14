import json
import sqlite3
from io import BytesIO
from decimal import Decimal
from datetime import datetime
from datetime import date
from datetime import timedelta
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import MagicMock, patch

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.base import ContentFile
from django.core.files.uploadedfile import SimpleUploadedFile
from django.core.management import call_command
from django.test import TestCase
from django.test.utils import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import Workbook, load_workbook

from .models import CompletedServiceAttachment, CompletedServiceEntry, ContractAttachment, ContractEntry, DocumentEntry, EquipmentLoan, EquipmentLoanPhoto, FuturaDigitalEntry, GoogleWorkspaceEmail, Insumo, PhoneExtension, Requisition, RequisitionBudget, RequisitionBudgetAttachment, RequisitionBudgetHistory, RequisitionUpdate, Starlink, Ticket, TicketAttendance, TicketAutoPauseReview, TicketFailureType, TicketPending, TicketUpdate, TipEntry


@override_settings(AUTHENTICATION_BACKENDS=['django.contrib.auth.backends.ModelBackend'])
class TicketAccessTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.normal_user = user_model.objects.create_user(
            username='usuario.comum',
            password='senha@123',
        )
        self.other_user = user_model.objects.create_user(
            username='outro.usuario',
            password='senha@123',
        )
        self.ti_user = user_model.objects.create_user(
            username='usuario.ti',
            password='senha@123',
        )
        self.other_ti_user = user_model.objects.create_user(
            username='outro.ti',
            password='senha@123',
        )
        self.fabiano_user = user_model.objects.create_user(
            username='fabiano.polone',
            password='senha@123',
        )
        ti_group, _ = Group.objects.get_or_create(name='TI')
        self.ti_user.groups.add(ti_group)
        self.other_ti_user.groups.add(ti_group)
        self.fabiano_user.groups.add(ti_group)

    def _ticket_create_token(self):
        response = self.client.get(reverse('chamados_new'))
        return response.context['ticket_create_token']

    def test_normal_user_creates_ticket_and_sees_own_only(self):
        self.client.login(username='usuario.comum', password='senha@123')
        token = self._ticket_create_token()
        with patch('chamados.views.whatsapp.notify_group_new_ticket') as mock_notify:
            self.client.post(
                reverse('chamados_new'),
                data={
                    'ticket_create_token': token,
                    'title': 'Notebook sem rede',
                    'description': 'Nao conecta na rede corporativa.',
                    'priority': Ticket.Priority.ALTA,
                },
            )
        self.assertEqual(Ticket.objects.count(), 1)
        ticket = Ticket.objects.first()
        self.assertEqual(ticket.created_by, self.normal_user)
        mock_notify.assert_called_once_with(ticket)

        Ticket.objects.create(
            title='Teste externo',
            description='Outro chamado',
            priority=Ticket.Priority.BAIXA,
            created_by=self.other_user,
        )

        response = self.client.get(reverse('chamados_list'))
        self.assertContains(response, 'Notebook sem rede')
        self.assertNotContains(response, 'Teste externo')

    def test_user_can_create_ticket_with_new_failure_type(self):
        self.client.login(username='usuario.comum', password='senha@123')
        token = self._ticket_create_token()

        with patch('chamados.views.whatsapp.notify_group_new_ticket'):
            response = self.client.post(
                reverse('chamados_new'),
                data={
                    'ticket_create_token': token,
                    'failure_type': '__new__',
                    'new_failure_type_name': 'Impressora fiscal',
                    'title': 'Impressora fiscal travada',
                    'description': 'Equipamento nao conclui a impressao.',
                    'priority': Ticket.Priority.MEDIA,
                },
            )

        self.assertRedirects(response, reverse('chamados_list'))
        failure_type = TicketFailureType.objects.get(name='Impressora fiscal')
        ticket = Ticket.objects.get(title='Impressora fiscal travada')
        self.assertEqual(ticket.failure_type, failure_type.name)
        self.assertEqual(ticket.get_failure_type_display(), 'Impressora fiscal')

        detail_response = self.client.get(reverse('chamados_detail', args=[ticket.id]))
        self.assertContains(detail_response, 'Impressora fiscal')

    def test_ticket_creation_still_succeeds_if_whatsapp_notification_fails(self):
        self.client.login(username='usuario.comum', password='senha@123')
        token = self._ticket_create_token()
        with patch('chamados.views.whatsapp.notify_group_new_ticket', side_effect=RuntimeError('falha wapi')):
            response = self.client.post(
                reverse('chamados_new'),
                data={
                    'ticket_create_token': token,
                    'title': 'Notebook sem rede',
                    'description': 'Nao conecta na rede corporativa.',
                    'priority': Ticket.Priority.ALTA,
                },
            )

        self.assertRedirects(response, reverse('chamados_list'))
        self.assertTrue(Ticket.objects.filter(title='Notebook sem rede').exists())

    def test_duplicate_ticket_create_submit_with_same_token_is_ignored(self):
        self.client.login(username='usuario.comum', password='senha@123')
        token = self._ticket_create_token()
        payload = {
            'ticket_create_token': token,
            'title': 'Clique duplicado',
            'description': 'Usuario clicou duas vezes no botao criar.',
            'priority': Ticket.Priority.MEDIA,
        }

        with patch('chamados.views.whatsapp.notify_group_new_ticket') as mock_notify:
            first_response = self.client.post(reverse('chamados_new'), data=payload)
            second_response = self.client.post(reverse('chamados_new'), data=payload)

        self.assertRedirects(first_response, reverse('chamados_list'))
        self.assertRedirects(second_response, reverse('chamados_list'))
        self.assertEqual(Ticket.objects.filter(title='Clique duplicado').count(), 1)
        self.assertEqual(mock_notify.call_count, 1)

    def test_normal_user_cannot_access_other_ticket(self):
        ticket = Ticket.objects.create(
            title='Problema de impressora',
            description='Falha ao imprimir.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.other_user,
        )
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_detail', args=[ticket.id]))
        self.assertRedirects(response, reverse('chamados_list'))

    def test_only_fabiano_can_delete_ticket(self):
        ticket = Ticket.objects.create(
            title='Chamado descartavel',
            description='Exclusao permitida apenas para o usuario definido.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
        )
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.post(reverse('chamados_delete', args=[ticket.id]), follow=True)
        self.assertContains(response, 'Voce nao possui permissao para excluir este chamado.')
        self.assertTrue(Ticket.objects.filter(id=ticket.id).exists())

        self.client.logout()
        self.client.login(username='fabiano.polone', password='senha@123')
        response = self.client.post(reverse('chamados_delete', args=[ticket.id]))
        self.assertRedirects(response, reverse('chamados_list'))
        self.assertFalse(Ticket.objects.filter(id=ticket.id).exists())

    def test_ti_can_export_attendances_to_spreadsheet(self):
        ticket = Ticket.objects.create(
            title='Planilha de teste',
            description='Falha ao acessar a impressora do financeiro.',
            priority=Ticket.Priority.ALTA,
            failure_type=Ticket.FailureType.HARDWARE,
            created_by=self.normal_user,
        )
        attendance = TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 4, 17, 8, 0)),
            ended_at=timezone.make_aware(datetime(2026, 4, 17, 9, 30)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Reinstalado driver e validado teste de impressao.',
        )

        workbook_buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Abril 2026'
        ws.append(['TI', 'Data', 'Contato', 'Setor', 'Notificacao', 'Prioridade', 'Falha', 'Acao / Correcao', 'Fechado', 'Tempo', 'Acao eficaz'])
        wb.save(workbook_buffer)
        workbook_buffer.seek(0)

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_preencher_planilha'),
            data={
                'attendant_id': self.ti_user.id,
                'export_month': '2026-04',
                'workbook_file': SimpleUploadedFile(
                    'chamados.xlsx',
                    workbook_buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
                'next': reverse('chamados_list'),
            },
        )

        self.assertEqual(response.status_code, 200)
        attendance.refresh_from_db()
        self.assertIsNotNone(attendance.exported_at)
        self.assertEqual(attendance.exported_path, 'upload:chamados.xlsx')

        saved = load_workbook(BytesIO(response.content))
        sheet = saved['Abril 2026']
        self.assertEqual(sheet.cell(row=2, column=1).value, ticket.id)
        self.assertEqual(sheet.cell(row=2, column=3).value, 'usuario.comum')
        self.assertEqual(sheet.cell(row=2, column=4).value, None)
        self.assertEqual(sheet.cell(row=2, column=5).value, 'Falha ao acessar a impressora do financeiro.')
        self.assertEqual(sheet.cell(row=2, column=6).value, 'Alta')
        self.assertEqual(sheet.cell(row=2, column=7).value, 'Hardware')
        self.assertEqual(sheet.cell(row=2, column=8).value, 'Reinstalado driver e validado teste de impressao.')
        self.assertEqual(sheet.cell(row=2, column=10).value, None)

    def test_spreadsheet_export_uses_ti_department_only_for_ti_ticket_creator(self):
        self.normal_user.email = 'usuario.comum@sidertec.com.br'
        self.normal_user.save(update_fields=['email'])
        normal_ticket = Ticket.objects.create(
            title='Chamado usuario comum',
            description='Nao deve preencher setor com dominio do email.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
        )
        ti_ticket = Ticket.objects.create(
            title='Chamado criado pela TI',
            description='Deve preencher setor como TI.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.ti_user,
        )
        TicketAttendance.objects.create(
            ticket=normal_ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 4, 17, 8, 0)),
            ended_at=timezone.make_aware(datetime(2026, 4, 17, 8, 30)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Atendimento usuario comum.',
        )
        TicketAttendance.objects.create(
            ticket=ti_ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 4, 17, 9, 0)),
            ended_at=timezone.make_aware(datetime(2026, 4, 17, 9, 30)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Atendimento chamado TI.',
        )

        workbook_buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Abril 2026'
        ws.append(['TI', 'Data', 'Contato', 'Setor', 'Notificacao', 'Prioridade', 'Falha', 'Acao / Correcao', 'Fechado', 'Tempo', 'Acao eficaz'])
        wb.save(workbook_buffer)
        workbook_buffer.seek(0)

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_preencher_planilha'),
            data={
                'attendant_id': self.ti_user.id,
                'export_month': '2026-04',
                'workbook_file': SimpleUploadedFile(
                    'chamados.xlsx',
                    workbook_buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
                'next': reverse('chamados_list'),
            },
        )

        self.assertEqual(response.status_code, 200)
        saved = load_workbook(BytesIO(response.content))
        sheet = saved['Abril 2026']
        self.assertEqual(sheet.cell(row=2, column=1).value, normal_ticket.id)
        self.assertEqual(sheet.cell(row=2, column=4).value, None)
        self.assertEqual(sheet.cell(row=3, column=1).value, ti_ticket.id)
        self.assertEqual(sheet.cell(row=3, column=4).value, 'TI')

    def test_ti_can_export_attendances_to_uploaded_spreadsheet(self):
        ticket = Ticket.objects.create(
            title='Planilha enviada',
            description='Chamado preenchido via arquivo selecionado.',
            priority=Ticket.Priority.MEDIA,
            failure_type=Ticket.FailureType.SOFTWARE,
            created_by=self.normal_user,
        )
        attendance = TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 4, 18, 10, 0)),
            ended_at=timezone.make_aware(datetime(2026, 4, 18, 11, 15)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Atendimento finalizado e validado.',
        )
        workbook_buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Abril 2026'
        ws.append(['TI', 'Data', 'Contato', 'Setor', 'Notificacao', 'Prioridade', 'Falha', 'Acao / Correcao', 'Fechado', 'Tempo', 'Acao eficaz'])
        wb.save(workbook_buffer)
        workbook_buffer.seek(0)

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_preencher_planilha'),
            data={
                'attendant_id': self.ti_user.id,
                'export_month': '2026-04',
                'workbook_file': SimpleUploadedFile(
                    'chamados.xlsx',
                    workbook_buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
                'next': reverse('chamados_list'),
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('preenchida-chamados.xlsx', response['Content-Disposition'])
        attendance.refresh_from_db()
        self.assertIsNotNone(attendance.exported_at)
        self.assertEqual(attendance.exported_path, 'upload:chamados.xlsx')

        saved = load_workbook(BytesIO(response.content))
        sheet = saved['Abril 2026']
        self.assertEqual(sheet.cell(row=2, column=1).value, ticket.id)
        self.assertEqual(sheet.cell(row=2, column=5).value, 'Chamado preenchido via arquivo selecionado.')
        self.assertEqual(sheet.cell(row=2, column=7).value, 'Software')
        self.assertEqual(sheet.cell(row=2, column=10).value, None)

    def test_spreadsheet_export_creates_one_row_per_attendance_cycle(self):
        ticket = Ticket.objects.create(
            title='Chamado com dois atendimentos',
            description='Cada play com pause ou stop deve virar uma linha.',
            priority=Ticket.Priority.ALTA,
            failure_type=Ticket.FailureType.HARDWARE,
            created_by=self.normal_user,
        )
        first_attendance = TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 4, 18, 8, 0)),
            ended_at=timezone.make_aware(datetime(2026, 4, 18, 8, 25)),
            end_action=TicketAttendance.EndAction.PAUSE,
            note='Primeira verificacao e pausa para aguardar usuario.',
        )
        second_attendance = TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 4, 18, 9, 0)),
            ended_at=timezone.make_aware(datetime(2026, 4, 18, 9, 40)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Retomado atendimento e finalizado.',
        )

        workbook_buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Abril 2026'
        ws.append(['TI', 'Data', 'Contato', 'Setor', 'Notificacao', 'Prioridade', 'Falha', 'Acao / Correcao', 'Fechado', 'Tempo', 'Acao eficaz'])
        wb.save(workbook_buffer)
        workbook_buffer.seek(0)

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_preencher_planilha'),
            data={
                'attendant_id': self.ti_user.id,
                'export_month': '2026-04',
                'workbook_file': SimpleUploadedFile(
                    'chamados.xlsx',
                    workbook_buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
                'next': reverse('chamados_list'),
            },
        )

        self.assertEqual(response.status_code, 200)
        first_attendance.refresh_from_db()
        second_attendance.refresh_from_db()
        self.assertIsNotNone(first_attendance.exported_at)
        self.assertIsNotNone(second_attendance.exported_at)

        saved = load_workbook(BytesIO(response.content))
        sheet = saved['Abril 2026']
        self.assertEqual(sheet.max_row, 3)
        self.assertEqual(sheet.cell(row=2, column=1).value, ticket.id)
        self.assertEqual(sheet.cell(row=2, column=2).value, '18/04/2026 08:00')
        self.assertEqual(sheet.cell(row=2, column=8).value, 'Primeira verificacao e pausa para aguardar usuario.')
        self.assertEqual(sheet.cell(row=2, column=9).value, '18/04/2026 08:25')
        self.assertEqual(sheet.cell(row=2, column=10).value, None)
        self.assertEqual(sheet.cell(row=3, column=1).value, ticket.id)
        self.assertEqual(sheet.cell(row=3, column=2).value, '18/04/2026 09:00')
        self.assertEqual(sheet.cell(row=3, column=8).value, 'Retomado atendimento e finalizado.')
        self.assertEqual(sheet.cell(row=3, column=9).value, '18/04/2026 09:40')
        self.assertEqual(sheet.cell(row=3, column=10).value, None)

    def test_spreadsheet_export_filters_by_selected_month(self):
        april_ticket = Ticket.objects.create(
            title='Chamado abril',
            description='Atendimento de abril deve entrar.',
            priority=Ticket.Priority.MEDIA,
            failure_type=Ticket.FailureType.SOFTWARE,
            created_by=self.normal_user,
        )
        may_ticket = Ticket.objects.create(
            title='Chamado maio',
            description='Atendimento de maio nao deve entrar ao exportar abril.',
            priority=Ticket.Priority.ALTA,
            failure_type=Ticket.FailureType.HARDWARE,
            created_by=self.normal_user,
        )
        TicketAttendance.objects.create(
            ticket=april_ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 4, 30, 16, 0)),
            ended_at=timezone.make_aware(datetime(2026, 4, 30, 16, 30)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Atendimento de abril.',
        )
        TicketAttendance.objects.create(
            ticket=may_ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 5, 1, 8, 0)),
            ended_at=timezone.make_aware(datetime(2026, 5, 1, 8, 30)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Atendimento de maio.',
        )

        workbook_buffer = BytesIO()
        wb = Workbook()
        april_sheet = wb.active
        april_sheet.title = 'Abril 2026'
        april_sheet.append(['TI', 'Data', 'Contato', 'Setor', 'Notificacao', 'Prioridade', 'Falha', 'Acao / Correcao', 'Fechado', 'Tempo', 'Acao eficaz'])
        may_sheet = wb.create_sheet('Maio 2026')
        may_sheet.append(['TI', 'Data', 'Contato', 'Setor', 'Notificacao', 'Prioridade', 'Falha', 'Acao / Correcao', 'Fechado', 'Tempo', 'Acao eficaz'])
        wb.save(workbook_buffer)
        workbook_buffer.seek(0)

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_preencher_planilha'),
            data={
                'attendant_id': self.ti_user.id,
                'export_month': '2026-04',
                'workbook_file': SimpleUploadedFile(
                    'chamados.xlsx',
                    workbook_buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
                'next': reverse('chamados_list'),
            },
        )

        self.assertEqual(response.status_code, 200)
        saved = load_workbook(BytesIO(response.content))
        self.assertEqual(saved['Abril 2026'].cell(row=2, column=1).value, april_ticket.id)
        self.assertEqual(saved['Maio 2026'].max_row, 1)

    def test_spreadsheet_export_orders_rows_by_started_at(self):
        later_ticket = Ticket.objects.create(
            title='Chamado mais tarde',
            description='Deve aparecer depois.',
            priority=Ticket.Priority.MEDIA,
            failure_type=Ticket.FailureType.SOFTWARE,
            created_by=self.normal_user,
        )
        earlier_ticket = Ticket.objects.create(
            title='Chamado mais cedo',
            description='Deve aparecer primeiro.',
            priority=Ticket.Priority.ALTA,
            failure_type=Ticket.FailureType.HARDWARE,
            created_by=self.normal_user,
        )
        TicketAttendance.objects.create(
            ticket=later_ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 4, 18, 14, 0)),
            ended_at=timezone.make_aware(datetime(2026, 4, 18, 14, 20)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Atendimento mais tarde.',
        )
        TicketAttendance.objects.create(
            ticket=earlier_ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 4, 18, 8, 0)),
            ended_at=timezone.make_aware(datetime(2026, 4, 18, 8, 20)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Atendimento mais cedo.',
        )

        workbook_buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Abril 2026'
        ws.append(['TI', 'Data', 'Contato', 'Setor', 'Notificacao', 'Prioridade', 'Falha', 'Acao / Correcao', 'Fechado', 'Tempo', 'Acao eficaz'])
        wb.save(workbook_buffer)
        workbook_buffer.seek(0)

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_preencher_planilha'),
            data={
                'attendant_id': self.ti_user.id,
                'export_month': '2026-04',
                'workbook_file': SimpleUploadedFile(
                    'chamados.xlsx',
                    workbook_buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
                'next': reverse('chamados_list'),
            },
        )

        self.assertEqual(response.status_code, 200)
        saved = load_workbook(BytesIO(response.content))
        sheet = saved['Abril 2026']
        self.assertEqual(sheet.cell(row=2, column=1).value, earlier_ticket.id)
        self.assertEqual(sheet.cell(row=2, column=2).value, '18/04/2026 08:00')
        self.assertEqual(sheet.cell(row=3, column=1).value, later_ticket.id)
        self.assertEqual(sheet.cell(row=3, column=2).value, '18/04/2026 14:00')

    def test_spreadsheet_export_compares_workbook_and_adds_only_missing_tickets(self):
        existing_ticket = Ticket.objects.create(
            title='Chamado ja na planilha',
            description='Ja existe na planilha.',
            priority=Ticket.Priority.ALTA,
            failure_type=Ticket.FailureType.HARDWARE,
            created_by=self.normal_user,
        )
        missing_ticket = Ticket.objects.create(
            title='Chamado novo para planilha',
            description='Deve entrar na planilha.',
            priority=Ticket.Priority.MEDIA,
            failure_type=Ticket.FailureType.SOFTWARE,
            created_by=self.normal_user,
        )
        exported_at = timezone.make_aware(datetime(2026, 5, 5, 12, 0))
        existing_attendance = TicketAttendance.objects.create(
            ticket=existing_ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 5, 5, 8, 0)),
            ended_at=timezone.make_aware(datetime(2026, 5, 5, 9, 15)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Nao deve duplicar.',
            exported_at=exported_at,
            exported_path='upload:antigo.xlsx',
        )
        missing_attendance = TicketAttendance.objects.create(
            ticket=missing_ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 5, 6, 10, 0)),
            ended_at=timezone.make_aware(datetime(2026, 5, 6, 11, 30)),
            end_action=TicketAttendance.EndAction.STOP,
            note='Deve ser exportado mesmo com exported_at preenchido.',
            exported_at=exported_at,
            exported_path='upload:antigo.xlsx',
        )

        workbook_buffer = BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Maio 2026'
        ws.append(['TI', 'Data', 'Contato', 'Setor', 'Notificacao', 'Prioridade', 'Falha', 'Acao / Correcao', 'Fechado', 'Tempo', 'Acao eficaz'])
        ws.append([existing_ticket.id, '05/05/2026 08:00', 'usuario.comum', '', 'Descricao antiga', 'Alta', 'Hardware', 'Acao antiga', '05/05/2026 09:15', '01:15', ''])
        wb.save(workbook_buffer)
        workbook_buffer.seek(0)

        self.client.login(username='usuario.ti', password='senha@123')
        with patch('chamados.excel_export.timezone.now', return_value=timezone.make_aware(datetime(2026, 5, 6, 10, 0))):
            response = self.client.post(
                reverse('chamados_preencher_planilha'),
                data={
                    'attendant_id': self.ti_user.id,
                    'export_month': '2026-05',
                    'workbook_file': SimpleUploadedFile(
                        'chamados.xlsx',
                        workbook_buffer.getvalue(),
                        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    ),
                    'next': reverse('chamados_list'),
                },
            )

        self.assertEqual(response.status_code, 200)
        existing_attendance.refresh_from_db()
        missing_attendance.refresh_from_db()
        self.assertEqual(existing_attendance.exported_at, exported_at)
        self.assertNotEqual(missing_attendance.exported_at, exported_at)
        self.assertEqual(missing_attendance.exported_path, 'upload:chamados.xlsx')

        saved = load_workbook(BytesIO(response.content))
        sheet = saved['Maio 2026']
        self.assertEqual(sheet.max_row, 3)
        self.assertEqual(sheet.cell(row=2, column=1).value, existing_ticket.id)
        self.assertEqual(sheet.cell(row=3, column=1).value, missing_ticket.id)
        self.assertEqual(sheet.cell(row=3, column=5).value, 'Deve entrar na planilha.')
        self.assertEqual(sheet.cell(row=3, column=7).value, 'Software')
        self.assertEqual(sheet.cell(row=3, column=8).value, 'Deve ser exportado mesmo com exported_at preenchido.')
        self.assertEqual(sheet.cell(row=3, column=10).value, None)

    def test_spreadsheet_export_is_blocked_when_auto_pause_review_is_pending(self):
        ticket = Ticket.objects.create(
            title='Chamado com pausa automatica',
            description='Descricao qualquer.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
        )
        attendance = TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=timezone.make_aware(datetime(2026, 4, 17, 8, 0)),
            ended_at=timezone.make_aware(datetime(2026, 4, 17, 9, 0)),
            end_action=TicketAttendance.EndAction.PAUSE,
            note='Atendimento encerrado automaticamente.',
        )
        TicketAutoPauseReview.objects.create(attendance=attendance)

        workbook_buffer = BytesIO()
        wb = Workbook()
        wb.save(workbook_buffer)
        workbook_buffer.seek(0)

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_preencher_planilha'),
            data={
                'attendant_id': self.ti_user.id,
                'export_month': '2026-04',
                'workbook_file': SimpleUploadedFile(
                    'chamados.xlsx',
                    workbook_buffer.getvalue(),
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ),
                'next': reverse('chamados_list'),
            },
            follow=True,
        )

        self.assertContains(response, 'Existem pausas automaticas pendentes para este atendente.')
        attendance.refresh_from_db()
        self.assertIsNone(attendance.exported_at)

    def test_ti_user_can_play_and_pause_ticket(self):
        ticket = Ticket.objects.create(
            title='VPN caiu',
            description='Sem acesso remoto.',
            priority=Ticket.Priority.CRITICA,
            created_by=self.normal_user,
        )
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(reverse('chamados_action', args=[ticket.id]), data={'action': 'play', 'next': reverse('chamados_list')})
        self.assertRedirects(response, reverse('chamados_list'))
        ticket.refresh_from_db()
        self.assertEqual(ticket.status, Ticket.Status.EM_ATENDIMENTO)
        running = TicketAttendance.objects.get(ticket=ticket, attendant=self.ti_user)
        self.assertIsNone(running.ended_at)

        response = self.client.post(
            reverse('chamados_action', args=[ticket.id]),
            data={'action': 'pause', 'next': reverse('chamados_list')},
            follow=True,
        )
        self.assertContains(response, 'Informe o que foi feito antes de pausar/parar.')
        running.refresh_from_db()
        self.assertIsNone(running.ended_at)

        response = self.client.post(
            reverse('chamados_action', args=[ticket.id]),
            data={
                'action': 'pause',
                'note': 'Rede estabilizada e usuario orientado.',
                'pause_status': Ticket.Status.ABERTO,
                'next': reverse('chamados_list'),
            },
        )
        self.assertRedirects(response, reverse('chamados_list'))
        ticket.refresh_from_db()
        running.refresh_from_db()
        self.assertEqual(ticket.status, Ticket.Status.ABERTO)
        self.assertIsNotNone(running.ended_at)
        self.assertEqual(running.end_action, TicketAttendance.EndAction.PAUSE)
        self.assertEqual(running.note, 'Rede estabilizada e usuario orientado.')

    def test_ti_can_pause_ticket_as_aguardando_usuario(self):
        ticket = Ticket.objects.create(
            title='Liberacao pendente do usuario',
            description='Aguardando retorno do usuario.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
            status=Ticket.Status.EM_ATENDIMENTO,
        )
        TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=ticket.created_at,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_action', args=[ticket.id]),
            data={
                'action': 'pause',
                'note': 'Solicitado retorno do usuario para teste final.',
                'pause_status': Ticket.Status.AGUARDANDO_USUARIO,
                'next': reverse('chamados_list'),
            },
        )

        self.assertRedirects(response, reverse('chamados_list'))
        ticket.refresh_from_db()
        self.assertEqual(ticket.status, Ticket.Status.AGUARDANDO_USUARIO)

    def test_ti_can_update_ticket_priority(self):
        ticket = Ticket.objects.create(
            title='Prioridade para alterar',
            description='Chamado aberto por usuario comum.',
            priority=Ticket.Priority.BAIXA,
            created_by=self.normal_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_action', args=[ticket.id]),
            data={
                'action': 'priority',
                'priority': Ticket.Priority.CRITICA,
                'next': reverse('chamados_detail', args=[ticket.id]),
            },
        )

        self.assertRedirects(response, reverse('chamados_detail', args=[ticket.id]))
        ticket.refresh_from_db()
        self.assertEqual(ticket.priority, Ticket.Priority.CRITICA)
        self.assertTrue(
            TicketUpdate.objects.filter(
                ticket=ticket,
                message__icontains='Prioridade alterada',
            ).exists()
        )

    def test_ti_can_stop_ticket_and_it_becomes_closed(self):
        ticket = Ticket.objects.create(
            title='Chamado para fechar',
            description='Fluxo de encerramento.',
            priority=Ticket.Priority.ALTA,
            created_by=self.normal_user,
            status=Ticket.Status.EM_ATENDIMENTO,
        )
        attendance = TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=ticket.created_at,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_action', args=[ticket.id]),
            data={
                'action': 'stop',
                'note': 'Equipamento ajustado e validado.',
                'failure_type': Ticket.FailureType.EQUIPAMENTO,
                'next': reverse('chamados_list'),
            },
        )

        self.assertRedirects(response, reverse('chamados_list'))
        ticket.refresh_from_db()
        attendance.refresh_from_db()
        self.assertEqual(ticket.status, Ticket.Status.FECHADO)
        self.assertEqual(ticket.failure_type, Ticket.FailureType.EQUIPAMENTO)
        self.assertIsNotNone(ticket.closed_at)
        self.assertEqual(attendance.end_action, TicketAttendance.EndAction.STOP)

    def test_ti_can_register_new_failure_type_when_stopping_ticket(self):
        ticket = Ticket.objects.create(
            title='Chamado para fechar com falha nova',
            description='Fluxo de encerramento com tipo novo.',
            priority=Ticket.Priority.ALTA,
            created_by=self.normal_user,
            status=Ticket.Status.EM_ATENDIMENTO,
        )
        TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=ticket.created_at,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_action', args=[ticket.id]),
            data={
                'action': 'stop',
                'note': 'Ajustada regra de rede.',
                'failure_type': '__new__',
                'new_failure_type_name': 'Rede interna',
                'next': reverse('chamados_detail', args=[ticket.id]),
            },
        )

        self.assertRedirects(response, reverse('chamados_detail', args=[ticket.id]))
        ticket.refresh_from_db()
        self.assertTrue(TicketFailureType.objects.filter(name='Rede interna').exists())
        self.assertEqual(ticket.failure_type, 'Rede interna')
        self.assertEqual(ticket.get_failure_type_display(), 'Rede interna')

    def test_management_command_auto_pauses_running_tickets(self):
        ticket = Ticket.objects.create(
            title='Chamado auto pause',
            description='Deve sair do play no fim do expediente.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
            status=Ticket.Status.EM_ATENDIMENTO,
        )
        attendance = TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=ticket.created_at,
        )

        fake_now = timezone.make_aware(datetime(2026, 4, 17, 21, 0, 0))
        with patch('chamados.management.commands.autopause_open_tickets.timezone.now', return_value=fake_now):
            call_command('autopause_open_tickets')

        ticket.refresh_from_db()
        attendance.refresh_from_db()
        review = TicketAutoPauseReview.objects.get(attendance=attendance)
        self.assertEqual(ticket.status, Ticket.Status.ABERTO)
        self.assertIsNotNone(attendance.ended_at)
        self.assertEqual(attendance.end_action, TicketAttendance.EndAction.PAUSE)
        self.assertIsNone(review.completed_at)
        self.assertTrue(
            TicketUpdate.objects.filter(
                ticket=ticket,
                message__icontains='Pause automatico no fim do expediente',
            ).exists()
        )

    def test_management_command_skips_before_end_of_day_without_force(self):
        ticket = Ticket.objects.create(
            title='Chamado ainda em expediente',
            description='Nao deve pausar antes das 17:45.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
            status=Ticket.Status.EM_ATENDIMENTO,
        )
        attendance = TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=ticket.created_at,
        )

        fake_now = timezone.make_aware(datetime(2026, 4, 17, 17, 44, 0))
        with patch('chamados.management.commands.autopause_open_tickets.timezone.now', return_value=fake_now):
            call_command('autopause_open_tickets')

        ticket.refresh_from_db()
        attendance.refresh_from_db()
        self.assertEqual(ticket.status, Ticket.Status.EM_ATENDIMENTO)
        self.assertIsNone(attendance.ended_at)
        self.assertFalse(TicketAutoPauseReview.objects.filter(attendance=attendance).exists())

    def test_ti_queue_shows_only_free_or_own_tickets_and_hides_closed(self):
        free_ticket = Ticket.objects.create(
            title='Chamado livre',
            description='Aguardando primeiro atendimento.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
        )
        own_ticket = Ticket.objects.create(
            title='Meu chamado TI',
            description='Atendimento do usuario.ti.',
            priority=Ticket.Priority.ALTA,
            created_by=self.normal_user,
        )
        locked_ticket = Ticket.objects.create(
            title='Chamado de outro TI',
            description='Este chamado ja foi iniciado por outro atendente.',
            priority=Ticket.Priority.BAIXA,
            created_by=self.normal_user,
        )
        TicketAttendance.objects.create(
            ticket=own_ticket,
            attendant=self.ti_user,
            started_at=own_ticket.created_at,
        )
        TicketAttendance.objects.create(
            ticket=locked_ticket,
            attendant=self.other_ti_user,
            started_at=locked_ticket.created_at,
        )
        closed_ticket = Ticket.objects.create(
            title='Chamado fechado',
            description='Nao deve aparecer na fila principal.',
            priority=Ticket.Priority.BAIXA,
            status=Ticket.Status.FECHADO,
            created_by=self.normal_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_list'))
        self.assertContains(response, free_ticket.title)
        self.assertContains(response, own_ticket.title)
        self.assertNotContains(response, locked_ticket.title)
        self.assertContains(response, f'Fechados (1)')
        self.assertContains(response, 'spreadsheetFileInput')
        self.assertContains(response, 'Planilha a ser exportada')
        self.assertContains(response, 'Exportar chamados novos')
        self.assertNotContains(response, 'spreadsheetPathInput')
        self.assertNotContains(response, 'Planilha no servidor/VPS')
        self.assertNotContains(response, 'refillCurrentMonthInput')
        self.assertNotContains(response, 'fillOriginalSpreadsheetButton')
        self.assertNotContains(response, closed_ticket.title)

        response = self.client.get(reverse('chamados_detail', args=[locked_ticket.id]))
        self.assertRedirects(response, reverse('chamados_list'))

        closed_response = self.client.get(reverse('chamados_closed_data'))
        self.assertEqual(closed_response.status_code, 200)
        self.assertIn(closed_ticket.title, closed_response.json()['items'][0]['title'])


    def test_ticket_list_highlights_non_ti_creator(self):
        Ticket.objects.create(
            title='Chamado criado por usuario comum',
            description='Nome do solicitante deve ficar destacado.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
        )
        Ticket.objects.create(
            title='Chamado criado por TI',
            description='Nome do TI nao deve usar a cor de solicitante externo.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_list'))

        self.assertContains(response, 'Chamado criado por usuario comum')
        self.assertContains(response, 'Chamado criado por TI')
        self.assertContains(response, 'class="ticket-creator-external"', count=1)

    def test_closed_tickets_data_filters_by_attendant_and_closed_date(self):
        first_closed = Ticket.objects.create(
            title='Fechado por usuario TI',
            description='Chamado filtrado por atendente e data.',
            priority=Ticket.Priority.MEDIA,
            status=Ticket.Status.FECHADO,
            created_by=self.normal_user,
        )
        second_closed = Ticket.objects.create(
            title='Fechado por outro TI',
            description='Nao deve entrar no filtro do usuario.ti.',
            priority=Ticket.Priority.MEDIA,
            status=Ticket.Status.FECHADO,
            created_by=self.normal_user,
        )
        first_closed_at = timezone.make_aware(datetime(2026, 5, 10, 9, 30))
        second_closed_at = timezone.make_aware(datetime(2026, 5, 11, 14, 0))
        Ticket.objects.filter(pk=first_closed.pk).update(closed_at=first_closed_at, updated_at=first_closed_at)
        Ticket.objects.filter(pk=second_closed.pk).update(closed_at=second_closed_at, updated_at=second_closed_at)
        TicketAttendance.objects.create(
            ticket=first_closed,
            attendant=self.ti_user,
            started_at=first_closed_at - timedelta(hours=1),
            ended_at=first_closed_at,
            end_action=TicketAttendance.EndAction.STOP,
        )
        TicketAttendance.objects.create(
            ticket=second_closed,
            attendant=self.other_ti_user,
            started_at=second_closed_at - timedelta(hours=1),
            ended_at=second_closed_at,
            end_action=TicketAttendance.EndAction.STOP,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(
            reverse('chamados_closed_data'),
            data={
                'attendant': 'usuario.ti',
                'date_from': '2026-05-10',
                'date_to': '2026-05-10',
            },
        )

        self.assertEqual(response.status_code, 200)
        items = response.json()['items']
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]['title'], first_closed.title)
        self.assertEqual(items[0]['attendant'], 'usuario.ti')
        self.assertIn('10/05/2026', items[0]['closed_at'])

    def test_ti_queue_includes_ticket_with_own_finished_attendance(self):
        reopened_like_ticket = Ticket.objects.create(
            title='Problemas com Microsoft Word',
            description='Historico de atendimento proprio, sem atendimento ativo.',
            priority=Ticket.Priority.MEDIA,
            status=Ticket.Status.ABERTO,
            created_by=self.normal_user,
        )
        TicketAttendance.objects.create(
            ticket=reopened_like_ticket,
            attendant=self.ti_user,
            started_at=reopened_like_ticket.created_at,
            ended_at=reopened_like_ticket.created_at,
            end_action=TicketAttendance.EndAction.PAUSE,
            note='Ciclo anterior finalizado.',
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_list'))
        self.assertContains(response, reopened_like_ticket.title)

    def test_ti_queue_hides_ticket_with_only_other_finished_attendance(self):
        hidden_ticket = Ticket.objects.create(
            title='Chamado pausado por outro atendente',
            description='Nao deve aparecer para quem nunca atendeu.',
            priority=Ticket.Priority.MEDIA,
            status=Ticket.Status.ABERTO,
            created_by=self.normal_user,
        )
        TicketAttendance.objects.create(
            ticket=hidden_ticket,
            attendant=self.other_ti_user,
            started_at=hidden_ticket.created_at,
            ended_at=hidden_ticket.created_at,
            end_action=TicketAttendance.EndAction.PAUSE,
            note='Atendimento anterior de outro atendente.',
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_list'))
        self.assertNotContains(response, hidden_ticket.title)

    def test_ti_can_consult_tickets_by_selected_attendant(self):
        free_ticket = Ticket.objects.create(
            title='Chamado livre geral',
            description='Sem atendente.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
        )
        own_ticket = Ticket.objects.create(
            title='Chamado do proprio atendente',
            description='Atendido pelo usuario.ti.',
            priority=Ticket.Priority.ALTA,
            created_by=self.normal_user,
        )
        other_ticket = Ticket.objects.create(
            title='Chamado do outro atendente',
            description='Atendido por outro.ti.',
            priority=Ticket.Priority.BAIXA,
            created_by=self.normal_user,
        )
        closed_other_ticket = Ticket.objects.create(
            title='Chamado fechado do outro atendente',
            description='Fechado por outro.ti e disponivel para consulta.',
            priority=Ticket.Priority.BAIXA,
            status=Ticket.Status.FECHADO,
            created_by=self.normal_user,
        )
        TicketAttendance.objects.create(
            ticket=own_ticket,
            attendant=self.ti_user,
            started_at=own_ticket.created_at,
        )
        TicketAttendance.objects.create(
            ticket=other_ticket,
            attendant=self.other_ti_user,
            started_at=other_ticket.created_at,
        )
        TicketAttendance.objects.create(
            ticket=closed_other_ticket,
            attendant=self.other_ti_user,
            started_at=closed_other_ticket.created_at,
            ended_at=closed_other_ticket.created_at,
            end_action=TicketAttendance.EndAction.STOP,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_list') + '?atendente=outro.ti')
        self.assertContains(response, 'Modo consulta ativo')
        self.assertContains(response, other_ticket.title)
        self.assertContains(response, closed_other_ticket.title)
        self.assertNotContains(response, free_ticket.title)
        self.assertNotContains(response, own_ticket.title)
        self.assertNotContains(response, '>usuario.ti<', html=False)

    def test_ti_can_claim_ticket_from_another_attendant_consultation(self):
        ticket = Ticket.objects.create(
            title='Chamado para puxar',
            description='Em atendimento com outro atendente.',
            priority=Ticket.Priority.MEDIA,
            status=Ticket.Status.EM_ATENDIMENTO,
            created_by=self.normal_user,
        )
        other_attendance = TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.other_ti_user,
            started_at=ticket.created_at,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_list') + '?atendente=outro.ti')
        self.assertContains(response, 'Puxar para mim')

        response = self.client.post(
            reverse('chamados_action', args=[ticket.id]),
            data={
                'action': 'claim',
                'next': reverse('chamados_list'),
            },
        )

        self.assertRedirects(response, reverse('chamados_list'))
        ticket.refresh_from_db()
        other_attendance.refresh_from_db()
        self.assertEqual(ticket.status, Ticket.Status.EM_ATENDIMENTO)
        self.assertIsNotNone(other_attendance.ended_at)
        self.assertEqual(other_attendance.end_action, TicketAttendance.EndAction.PAUSE)
        self.assertEqual(other_attendance.note, 'Transferido para usuario.ti.')
        self.assertTrue(
            TicketAttendance.objects.filter(
                ticket=ticket,
                attendant=self.ti_user,
                ended_at__isnull=True,
            ).exists()
        )
        self.assertTrue(
            TicketUpdate.objects.filter(
                ticket=ticket,
                message__icontains='Chamado puxado de outro.ti para usuario.ti',
            ).exists()
        )

    def test_ti_can_view_closed_ticket_from_another_attendant_without_consult_mode(self):
        closed_ticket = Ticket.objects.create(
            title='Chamado fechado por outro TI',
            description='Detalhe deve abrir para qualquer atendente TI.',
            priority=Ticket.Priority.MEDIA,
            status=Ticket.Status.FECHADO,
            created_by=self.normal_user,
        )
        TicketAttendance.objects.create(
            ticket=closed_ticket,
            attendant=self.other_ti_user,
            started_at=closed_ticket.created_at,
            ended_at=closed_ticket.created_at,
            end_action=TicketAttendance.EndAction.STOP,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_detail', args=[closed_ticket.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, closed_ticket.title)
        self.assertNotContains(response, 'Atendimento TI')

    def test_ti_can_review_auto_paused_tickets(self):
        ticket = Ticket.objects.create(
            title='Chamado revisao auto pause',
            description='Registro do dia seguinte.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
            status=Ticket.Status.EM_ATENDIMENTO,
        )
        attendance = TicketAttendance.objects.create(
            ticket=ticket,
            attendant=self.ti_user,
            started_at=ticket.created_at,
            ended_at=ticket.created_at,
            end_action=TicketAttendance.EndAction.PAUSE,
            note='',
        )
        review = TicketAutoPauseReview.objects.create(attendance=attendance)

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_auto_pause_reviews'))
        self.assertContains(response, 'Pausas automaticas')
        self.assertContains(response, ticket.title)

        save_response = self.client.post(
            reverse('chamados_auto_pause_reviews'),
            data={
                'review_id': review.id,
                'note': 'Troca concluida e validada antes de encerrar o expediente.',
                'status': Ticket.Status.FECHADO,
            },
        )
        self.assertRedirects(save_response, reverse('chamados_auto_pause_reviews'))

        ticket.refresh_from_db()
        attendance.refresh_from_db()
        review.refresh_from_db()
        self.assertEqual(ticket.status, Ticket.Status.FECHADO)
        self.assertEqual(attendance.note, 'Troca concluida e validada antes de encerrar o expediente.')
        self.assertIsNotNone(review.completed_at)

    def test_ti_can_open_other_attendant_ticket_in_consult_mode_read_only(self):
        locked_ticket = Ticket.objects.create(
            title='Chamado consulta',
            description='Somente leitura para outros atendentes.',
            priority=Ticket.Priority.MEDIA,
            created_by=self.normal_user,
        )
        TicketAttendance.objects.create(
            ticket=locked_ticket,
            attendant=self.other_ti_user,
            started_at=locked_ticket.created_at,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_detail', args=[locked_ticket.id]) + '?consult=1')
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Modo consulta')
        self.assertNotContains(response, 'Atendimento TI')

    def test_ticket_detail_hides_legacy_metadata_from_description_and_history(self):
        ticket = Ticket.objects.create(
            title='Chamado legado',
            description='Descricao util\n\nTipo legado: requisicao | Falha legado: -\n[ERP-TI-ID:343]',
            priority=Ticket.Priority.MEDIA,
            status=Ticket.Status.EM_ATENDIMENTO,
            created_by=self.normal_user,
        )
        TicketUpdate.objects.create(
            ticket=ticket,
            author=self.ti_user,
            message='Evento legado (assigned): novo -> em_atendimento\nChamado assumido por usuario.ti.\n[ERP-TI-EVENT:1552]',
            status_to=Ticket.Status.EM_ATENDIMENTO,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_detail', args=[ticket.id]))
        self.assertContains(response, 'Descricao util')
        self.assertContains(response, 'Chamado assumido por usuario.ti.')
        self.assertNotContains(response, 'Tipo legado:')
        self.assertNotContains(response, 'Falha legado:')
        self.assertNotContains(response, 'ERP-TI-ID')
        self.assertNotContains(response, 'Evento legado')
        self.assertNotContains(response, 'ERP-TI-EVENT')

    def test_only_ti_can_access_pending_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_pending_list'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_pending_list'))
        self.assertEqual(response.status_code, 200)

    def test_ti_pending_is_individual_and_can_be_deleted(self):
        own_pending = TicketPending.objects.create(
            attendant=self.ti_user,
            content='Revisar backup do servidor legado.',
        )
        TicketPending.objects.create(
            attendant=self.other_ti_user,
            content='Validar impressora do setor comercial.',
        )
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_pending_list'))
        self.assertContains(response, own_pending.content)
        self.assertNotContains(response, 'Validar impressora do setor comercial.')

        delete_response = self.client.post(reverse('chamados_pending_delete', args=[own_pending.id]))
        self.assertRedirects(delete_response, reverse('chamados_pending_list'))
        self.assertFalse(TicketPending.objects.filter(id=own_pending.id).exists())

    def test_create_ticket_from_pending_starts_attendance_with_programmed_priority(self):
        pending = TicketPending.objects.create(
            attendant=self.ti_user,
            content='Atualizar permissoes de acesso da pasta financeira.',
        )
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(reverse('chamados_pending_create_ticket', args=[pending.id]))
        self.assertRedirects(response, reverse('chamados_list'))

        ticket = Ticket.objects.get()
        self.assertEqual(ticket.created_by, self.ti_user)
        self.assertEqual(ticket.priority, Ticket.Priority.PROGRAMADA)
        self.assertEqual(ticket.status, Ticket.Status.EM_ATENDIMENTO)
        self.assertEqual(ticket.title, 'Atualizar permissoes de acesso da pasta financeira.')
        self.assertIn('Atualizar permissoes de acesso da pasta financeira.', ticket.description)

        running = TicketAttendance.objects.get(ticket=ticket, attendant=self.ti_user)
        self.assertIsNone(running.ended_at)
        self.assertFalse(TicketPending.objects.filter(id=pending.id).exists())

    def test_only_ti_can_access_requisicoes_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Requisicoes TI')

    def test_ti_can_create_and_edit_requisition(self):
        self.client.login(username='usuario.ti', password='senha@123')
        payload_create = json.dumps(
            [
                {
                    'id': '',
                    'temp_key': 'tmp_root_1',
                    'parent_ref': '',
                    'store_name': 'Kabum',
                    'title': 'Orcamento principal',
                    'amount': '1500.00',
                    'quantity': '2',
                    'freight_amount': '150.00',
                    'discount_amount': '100.00',
                    'approval_status': RequisitionBudget.ApprovalStatus.APROVADO,
                    'receipt_status': RequisitionBudget.ReceiptStatus.PARCIAL,
                    'received_quantity': '1',
                    'notes': 'Fornecedor A',
                    'file_key': 'budget_file_tmp_root_1',
                    'clear_file': False,
                },
                {
                    'id': '',
                    'temp_key': 'tmp_sub_1',
                    'parent_ref': 'tmp:tmp_root_1',
                    'store_name': 'Instaladora XPTO',
                    'title': 'Suborcamento de instalacao',
                    'amount': '300.00',
                    'quantity': '3',
                    'freight_amount': '30.00',
                    'discount_amount': '0',
                    'approval_status': RequisitionBudget.ApprovalStatus.PENDENTE,
                    'receipt_status': RequisitionBudget.ReceiptStatus.PENDENTE,
                    'received_quantity': '0',
                    'notes': '',
                    'file_key': 'budget_file_tmp_sub_1',
                    'clear_file': False,
                },
            ]
        )
        create_response = self.client.post(
            reverse('chamados_requisicoes_save'),
            data={
                'title': 'Compra de notebook para diretoria',
                'kind': Requisition.Kind.FISICA,
                'request_text': 'Necessario para substituicao do equipamento atual.',
                'budgets_payload': payload_create,
            },
        )
        self.assertRedirects(create_response, reverse('chamados_requisicoes'))
        requisition = Requisition.objects.get()
        self.assertEqual(requisition.requested_by, self.ti_user)
        self.assertEqual(requisition.status, Requisition.Status.APROVADA)
        self.assertTrue(requisition.code.startswith('REQ-'))
        self.assertEqual(RequisitionBudget.objects.filter(requisition=requisition).count(), 2)
        root_budget = RequisitionBudget.objects.get(requisition=requisition, parent_budget__isnull=True)
        sub_budget = RequisitionBudget.objects.get(requisition=requisition, parent_budget__isnull=False)
        self.assertEqual(sub_budget.parent_budget_id, root_budget.id)
        self.assertEqual(root_budget.store_name, 'Kabum')
        self.assertEqual(sub_budget.store_name, 'Instaladora XPTO')
        self.assertEqual(root_budget.quantity, 2)
        self.assertEqual(sub_budget.quantity, 3)
        self.assertEqual(str(root_budget.freight_amount), '150.00')
        self.assertEqual(str(sub_budget.freight_amount), '30.00')
        self.assertEqual(str(root_budget.discount_amount), '100.00')
        self.assertEqual(root_budget.approval_status, RequisitionBudget.ApprovalStatus.APROVADO)
        self.assertEqual(root_budget.receipt_status, RequisitionBudget.ReceiptStatus.PARCIAL)
        self.assertEqual(root_budget.received_quantity, 1)
        self.assertEqual(requisition.budget_total, Decimal('3980.00'))
        self.assertEqual(RequisitionBudgetHistory.objects.filter(budget=root_budget).count(), 1)

        payload_edit = json.dumps(
            [
                {
                    'id': str(root_budget.id),
                    'temp_key': 'tmp_root_1',
                    'parent_ref': '',
                    'store_name': 'Pichau',
                    'title': 'Orcamento principal atualizado',
                    'amount': '2000.00',
                    'quantity': '4',
                    'freight_amount': '89.90',
                    'discount_amount': '200.00',
                    'approval_status': RequisitionBudget.ApprovalStatus.APROVADO,
                    'receipt_status': RequisitionBudget.ReceiptStatus.RECEBIDO,
                    'received_quantity': '4',
                    'notes': 'Fornecedor B',
                    'file_key': 'budget_file_tmp_root_1',
                    'clear_file': False,
                }
            ]
        )
        edit_response = self.client.post(
            reverse('chamados_requisicoes_save'),
            data={
                'requisition_id': requisition.id,
                'title': 'Compra de notebook para presidencia',
                'kind': Requisition.Kind.FISICA,
                'request_text': 'Atualizacao da requisicao com especificacao de memoria.',
                'budgets_payload': payload_edit,
            },
        )
        self.assertRedirects(edit_response, reverse('chamados_requisicoes'))
        requisition.refresh_from_db()
        self.assertEqual(requisition.title, 'Compra de notebook para presidencia')
        self.assertEqual(RequisitionUpdate.objects.filter(requisition=requisition).count(), 3)
        self.assertEqual(RequisitionBudget.objects.filter(requisition=requisition).count(), 1)
        root_budget.refresh_from_db()
        self.assertEqual(root_budget.store_name, 'Pichau')
        self.assertEqual(root_budget.title, 'Orcamento principal atualizado')
        self.assertEqual(str(root_budget.amount), '2000.00')
        self.assertEqual(root_budget.quantity, 4)
        self.assertEqual(str(root_budget.freight_amount), '89.90')
        self.assertEqual(str(root_budget.discount_amount), '200.00')
        self.assertEqual(root_budget.receipt_status, RequisitionBudget.ReceiptStatus.RECEBIDO)
        self.assertEqual(root_budget.received_quantity, 4)
        self.assertEqual(requisition.budget_total, Decimal('7889.90'))
        self.assertEqual(RequisitionBudgetHistory.objects.filter(budget=root_budget).count(), 2)

    def test_requisition_save_auto_approves_when_any_budget_is_approved(self):
        self.client.login(username='usuario.ti', password='senha@123')
        payload = json.dumps(
            [
                {
                    'id': '',
                    'temp_key': 'tmp_root_1',
                    'parent_ref': '',
                    'store_name': 'Loja Exemplo',
                    'title': 'Notebook',
                    'amount': '2500.00',
                    'quantity': '1',
                    'discount_amount': '0',
                    'approval_status': RequisitionBudget.ApprovalStatus.APROVADO,
                    'receipt_status': RequisitionBudget.ReceiptStatus.PENDENTE,
                    'received_quantity': '0',
                    'notes': '',
                    'file_key': 'budget_file_tmp_root_1',
                    'clear_file': False,
                }
            ]
        )
        response = self.client.post(
            reverse('chamados_requisicoes_save'),
            data={
                'title': 'Compra emergencial',
                'kind': Requisition.Kind.FISICA,
                'request_text': 'Reposicao.',
                'budgets_payload': payload,
            },
        )

        self.assertRedirects(response, reverse('chamados_requisicoes'))
        requisition = Requisition.objects.get(title='Compra emergencial')
        self.assertEqual(requisition.status, Requisition.Status.APROVADA)

    def test_ti_can_update_requisition_status(self):
        requisition = Requisition.objects.create(
            title='Licenca de software de design',
            kind=Requisition.Kind.DIGITAL,
            request_text='Aquisicao anual para equipe de marketing.',
            requested_by=self.ti_user,
        )
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_requisicoes_status', args=[requisition.id]),
            data={
                'status': Requisition.Status.APROVADA,
                'note': 'Aprovado em reuniao mensal.',
            },
        )
        self.assertRedirects(response, reverse('chamados_requisicoes'))
        requisition.refresh_from_db()
        self.assertEqual(requisition.status, Requisition.Status.APROVADA)
        self.assertIsNotNone(requisition.approved_at)
        self.assertTrue(
            RequisitionUpdate.objects.filter(
                requisition=requisition,
                status_to=Requisition.Status.APROVADA,
            ).exists()
        )

    def test_ti_can_mark_approved_requisition_as_delivered_with_date(self):
        requisition = Requisition.objects.create(
            title='Compra de headset aprovada',
            kind=Requisition.Kind.FISICA,
            request_text='Compra ja aprovada para equipe comercial.',
            requested_by=self.ti_user,
            status=Requisition.Status.APROVADA,
            approved_at=date(2026, 5, 1),
        )
        budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor Headset',
            title='Headset USB',
            amount='180.00',
            quantity=3,
            approval_status=RequisitionBudget.ApprovalStatus.APROVADO,
            receipt_status=RequisitionBudget.ReceiptStatus.PENDENTE,
            received_quantity=0,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_requisicoes_deliver', args=[requisition.id]),
            data={
                'delivered_at': '2026-05-04',
                'note': 'Recebido pelo Fabiano.',
            },
        )

        self.assertRedirects(response, reverse('chamados_requisicoes'))
        requisition.refresh_from_db()
        budget.refresh_from_db()
        self.assertEqual(requisition.status, Requisition.Status.ENTREGUE)
        self.assertEqual(requisition.received_at, date(2026, 5, 4))
        self.assertEqual(budget.receipt_status, RequisitionBudget.ReceiptStatus.RECEBIDO)
        self.assertEqual(budget.received_quantity, 3)
        self.assertTrue(
            RequisitionUpdate.objects.filter(
                requisition=requisition,
                status_to=Requisition.Status.ENTREGUE,
                message__icontains='Compra entregue em 04/05/2026',
            ).exists()
        )
        self.assertTrue(
            RequisitionBudgetHistory.objects.filter(
                budget=budget,
                message__icontains='Compra entregue em 04/05/2026',
            ).exists()
        )

    def test_requisition_details_payload_includes_delivery_action_and_history(self):
        requisition = Requisition.objects.create(
            title='Compra aprovada com entrega',
            kind=Requisition.Kind.FISICA,
            request_text='Validar botao de entrega.',
            requested_by=self.ti_user,
            status=Requisition.Status.APROVADA,
            approved_at=date(2026, 5, 2),
        )
        RequisitionUpdate.objects.create(
            requisition=requisition,
            author=self.ti_user,
            message='Requisicao aprovada para compra.',
            status_to=Requisition.Status.APROVADA,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))

        payload = response.context['requisitions_payload'][0]
        self.assertTrue(payload['can_mark_delivered'])
        self.assertEqual(payload['deliver_url'], reverse('chamados_requisicoes_deliver', args=[requisition.id]))
        self.assertEqual(payload['approved_at_display'], '02/05/2026')
        self.assertEqual(payload['updates'][0]['message'], 'Requisicao aprovada para compra.')
        self.assertContains(response, 'requisitionDeliverForm')
        self.assertContains(response, 'requisitionDetailsHistory')

    def test_ti_can_reject_pending_requisition_and_all_budgets(self):
        requisition = Requisition.objects.create(
            title='Compra nao aprovada',
            kind=Requisition.Kind.FISICA,
            request_text='Nenhum orcamento aprovado.',
            requested_by=self.ti_user,
            status=Requisition.Status.PENDENTE_APROVACAO,
        )
        root_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor principal',
            title='Desktop',
            amount='2500.00',
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )
        sub_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            parent_budget=root_budget,
            store_name='Fornecedor sub',
            title='Memoria adicional',
            amount='300.00',
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(reverse('chamados_requisicoes_reject_all_budgets', args=[requisition.id]))

        self.assertRedirects(response, reverse('chamados_requisicoes'))
        requisition.refresh_from_db()
        root_budget.refresh_from_db()
        sub_budget.refresh_from_db()
        self.assertEqual(requisition.status, Requisition.Status.NAO_APROVADA)
        self.assertIsNone(requisition.approved_at)
        self.assertEqual(root_budget.approval_status, RequisitionBudget.ApprovalStatus.NAO_APROVADO)
        self.assertEqual(sub_budget.approval_status, RequisitionBudget.ApprovalStatus.NAO_APROVADO)
        self.assertTrue(
            RequisitionUpdate.objects.filter(
                requisition=requisition,
                status_to=Requisition.Status.NAO_APROVADA,
                message__icontains='2 orcamento(s) marcado(s) como nao aprovado(s)',
            ).exists()
        )
        self.assertEqual(
            RequisitionBudgetHistory.objects.filter(
                budget__in=[root_budget, sub_budget],
                message__icontains='rejeicao da requisicao',
            ).count(),
            2,
        )

    def test_ti_can_reapply_reject_all_on_already_rejected_requisition(self):
        requisition = Requisition.objects.create(
            title='Compra antiga nao aprovada',
            kind=Requisition.Kind.FISICA,
            request_text='Corrigir orcamentos antigos.',
            requested_by=self.ti_user,
            status=Requisition.Status.NAO_APROVADA,
        )
        budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor antigo',
            title='Item antigo',
            amount='180.00',
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(reverse('chamados_requisicoes_reject_all_budgets', args=[requisition.id]))

        self.assertRedirects(response, reverse('chamados_requisicoes'))
        budget.refresh_from_db()
        requisition.refresh_from_db()
        self.assertEqual(requisition.status, Requisition.Status.NAO_APROVADA)
        self.assertEqual(budget.approval_status, RequisitionBudget.ApprovalStatus.NAO_APROVADO)

    def test_requisition_payload_includes_reject_all_url(self):
        requisition = Requisition.objects.create(
            title='Compra aguardando decisao',
            kind=Requisition.Kind.FISICA,
            request_text='Validar botao nao aprovado.',
            requested_by=self.ti_user,
            status=Requisition.Status.PENDENTE_APROVACAO,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor pendente',
            title='Switch',
            amount='700.00',
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))

        payload = response.context['requisitions_payload'][0]
        self.assertEqual(
            payload['reject_all_url'],
            reverse('chamados_requisicoes_reject_all_budgets', args=[requisition.id]),
        )
        self.assertContains(response, 'requisitionRejectAllForm')

    def test_ti_can_approve_specific_requisition_budget(self):
        requisition = Requisition.objects.create(
            title='Compra de nobreak',
            kind=Requisition.Kind.FISICA,
            request_text='Reposicao do CPD.',
            requested_by=self.ti_user,
        )
        budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor Z',
            title='Nobreak 1500VA',
            amount='1800.00',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )
        sub_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            parent_budget=budget,
            store_name='Fornecedor Z',
            title='Modulo de bateria',
            amount='450.00',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(reverse('chamados_requisicoes_budget_approve', args=[budget.id]))

        self.assertRedirects(response, reverse('chamados_requisicoes'))
        budget.refresh_from_db()
        sub_budget.refresh_from_db()
        requisition.refresh_from_db()
        self.assertEqual(budget.approval_status, RequisitionBudget.ApprovalStatus.APROVADO)
        self.assertEqual(sub_budget.approval_status, RequisitionBudget.ApprovalStatus.APROVADO)
        self.assertEqual(requisition.status, Requisition.Status.APROVADA)
        self.assertTrue(
            RequisitionBudgetHistory.objects.filter(
                budget=budget,
                message__icontains='Orcamento aprovado diretamente pela visualizacao',
            ).exists()
        )
        self.assertTrue(
            RequisitionBudgetHistory.objects.filter(
                budget=sub_budget,
                message__icontains='Orcamento aprovado diretamente pela visualizacao',
            ).exists()
        )
        self.assertTrue(
            RequisitionUpdate.objects.filter(
                requisition=requisition,
                status_to=Requisition.Status.APROVADA,
                message__icontains='Requisicao aprovada a partir do orcamento',
            ).exists()
        )

    def test_approving_main_same_store_budget_approves_related_root_budgets(self):
        requisition = Requisition.objects.create(
            title='Pedido tablet',
            kind=Requisition.Kind.FISICA,
            request_text='Pedido com acessorios.',
            requested_by=self.ti_user,
        )
        main_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='MercadoLivre',
            title='Tablet',
            amount='1799.99',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )
        case_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='MercadoLivre',
            title='Capa',
            amount='159.20',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )
        film_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='MercadoLivre',
            title='Peliculas',
            amount='22.29',
            quantity=2,
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(reverse('chamados_requisicoes_budget_approve', args=[main_budget.id]))

        self.assertRedirects(response, reverse('chamados_requisicoes'))
        for budget in [main_budget, case_budget, film_budget]:
            budget.refresh_from_db()
            self.assertEqual(budget.approval_status, RequisitionBudget.ApprovalStatus.APROVADO)

    def test_ti_can_disapprove_specific_requisition_budget(self):
        requisition = Requisition.objects.create(
            title='Compra de monitor aprovado',
            kind=Requisition.Kind.FISICA,
            request_text='Compra aprovada por engano.',
            status=Requisition.Status.APROVADA,
            requested_by=self.ti_user,
            approved_at=date(2026, 4, 1),
        )
        budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor A',
            title='Monitor 24',
            amount='900.00',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.APROVADO,
        )
        sub_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            parent_budget=budget,
            store_name='Fornecedor A',
            title='Cabo HDMI',
            amount='40.00',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.APROVADO,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(reverse('chamados_requisicoes_budget_disapprove', args=[budget.id]))

        self.assertRedirects(response, reverse('chamados_requisicoes'))
        budget.refresh_from_db()
        sub_budget.refresh_from_db()
        requisition.refresh_from_db()
        self.assertEqual(budget.approval_status, RequisitionBudget.ApprovalStatus.NAO_APROVADO)
        self.assertEqual(sub_budget.approval_status, RequisitionBudget.ApprovalStatus.NAO_APROVADO)
        self.assertEqual(requisition.status, Requisition.Status.PENDENTE_APROVACAO)
        self.assertIsNone(requisition.approved_at)
        self.assertTrue(
            RequisitionBudgetHistory.objects.filter(
                budget=budget,
                message__icontains='Orcamento desaprovado diretamente pela visualizacao',
            ).exists()
        )
        self.assertTrue(
            RequisitionBudgetHistory.objects.filter(
                budget=sub_budget,
                message__icontains='Orcamento desaprovado diretamente pela visualizacao',
            ).exists()
        )

    def test_requisition_payload_marks_approved_budget_as_disapprovable(self):
        requisition = Requisition.objects.create(
            title='Compra com botao desaprovar',
            kind=Requisition.Kind.FISICA,
            request_text='Validar botao.',
            requested_by=self.ti_user,
        )
        budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor B',
            title='Notebook',
            amount='3500.00',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.APROVADO,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))

        payload = response.context['requisitions_payload'][0]['budgets'][0]
        self.assertFalse(payload['can_approve'])
        self.assertTrue(payload['can_disapprove'])
        self.assertEqual(payload['disapprove_url'], reverse('chamados_requisicoes_budget_disapprove', args=[budget.id]))

    def test_disapprove_button_does_not_skip_budget_attachment_rendering(self):
        template_path = Path(__file__).resolve().parents[1] / 'templates' / 'chamados' / 'requisicoes.html'
        content = template_path.read_text(encoding='utf-8')
        disapprove_index = content.index('Desaprovar orçamento')
        evidence_index = content.index('if (budget.evidence_url && budget.evidence_is_image)', disapprove_index)
        self.assertNotIn('return item;', content[disapprove_index:evidence_index])

    def test_rejected_budget_summary_strikes_budget_title(self):
        requisition = Requisition.objects.create(
            title='Compra rejeitada',
            kind=Requisition.Kind.FISICA,
            request_text='Validar destaque visual.',
            requested_by=self.ti_user,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor rejeitado',
            title='Item recusado',
            amount='120.00',
            approval_status=RequisitionBudget.ApprovalStatus.NAO_APROVADO,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))

        self.assertContains(response, 'requisition-budget-chip rejected')
        self.assertContains(response, '<span class="requisition-budget-chip-title">Fornecedor rejeitado</span>', html=True)
        css_path = Path(__file__).resolve().parents[1] / 'static' / 'css' / 'login.css'
        css_content = css_path.read_text(encoding='utf-8')
        self.assertIn('.requisition-budget-chip.rejected .requisition-budget-chip-title', css_content)
        self.assertIn('text-decoration: line-through;', css_content)

    def test_requisicoes_page_orders_newest_requisitions_first(self):
        older_requisition = Requisition.objects.create(
            title='Requisicao antiga',
            kind=Requisition.Kind.FISICA,
            request_text='Criada antes.',
            requested_by=self.ti_user,
        )
        newer_requisition = Requisition.objects.create(
            title='Requisicao nova',
            kind=Requisition.Kind.FISICA,
            request_text='Criada depois.',
            requested_by=self.ti_user,
        )
        now = timezone.now()
        Requisition.objects.filter(pk=older_requisition.pk).update(created_at=now - timedelta(days=2))
        Requisition.objects.filter(pk=newer_requisition.pk).update(created_at=now - timedelta(days=1))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))

        titles = [row['requisition'].title for row in response.context['requisition_rows']]
        self.assertEqual(titles, ['Requisicao nova', 'Requisicao antiga'])

    def test_requisicoes_page_shows_approval_date_in_list(self):
        Requisition.objects.create(
            title='Compra aprovada com data',
            kind=Requisition.Kind.FISICA,
            request_text='Mostrar data na listagem.',
            requested_by=self.ti_user,
            status=Requisition.Status.APROVADA,
            approved_at=date(2026, 4, 27),
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))

        self.assertContains(response, 'Aprovada em 27/04/2026')

    def test_requisicoes_page_reconciles_old_pending_status_when_budget_is_approved(self):
        requisition = Requisition.objects.create(
            title='Orcamento legado aprovado',
            kind=Requisition.Kind.FISICA,
            request_text='Registro antigo antes da sincronizacao automatica.',
            requested_by=self.ti_user,
            status=Requisition.Status.PENDENTE_APROVACAO,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor legado',
            title='Bateria de nobreak',
            amount='564.50',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.APROVADO,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))

        self.assertEqual(response.status_code, 200)
        requisition.refresh_from_db()
        self.assertEqual(requisition.status, Requisition.Status.APROVADA)
        self.assertContains(response, 'Aprovada')

    def test_requisicoes_page_has_copy_buttons(self):
        requisition = Requisition.objects.create(
            title='Compra de cadeira ergonomica',
            kind=Requisition.Kind.FISICA,
            request_text='Apoio para colaborador com recomendacao medica.',
            requested_by=self.ti_user,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor C',
            title='Orcamento principal',
            amount='980.00',
            quantity=2,
            discount_amount='30.00',
            notes='Fornecedor C',
        )
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))
        self.assertContains(response, 'Copiar para Email')
        self.assertContains(response, 'Copiar para WhatsApp')
        self.assertContains(response, 'Copiar relatório do mês')
        self.assertContains(response, '<span class="requisition-budget-chip-title">Fornecedor C</span>', html=True)
        self.assertContains(response, 'Qtd 2')
        self.assertContains(response, 'Unit. R$ 980,00')
        self.assertContains(response, 'R$ 1.930,00')
        self.assertContains(response, 'Pendente')
        self.assertNotContains(response, 'https://wa.me/')
        share_text = response.context['requisition_share_map'][str(requisition.id)]
        self.assertIn('------------------------------', share_text)
        self.assertIn('Orçamento 1', share_text)
        self.assertIn('Valor final: R$ 1.930,00', share_text)
        self.assertNotIn('Total geral', share_text)
        self.assertNotIn('Aprovação:', share_text)

    def test_requisition_copy_text_includes_sub_budget_totals(self):
        requisition = Requisition.objects.create(
            title='Compra com acessorios',
            kind=Requisition.Kind.FISICA,
            request_text='Compra com itens relacionados.',
            requested_by=self.ti_user,
        )
        root_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor Kit',
            title='Equipamento principal',
            amount='1000.00',
            quantity=1,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            parent_budget=root_budget,
            store_name='Fornecedor Kit',
            title='Licenca adicional',
            amount='150.00',
            quantity=2,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            parent_budget=root_budget,
            store_name='Fornecedor Kit',
            title='Instalacao',
            amount='250.00',
            quantity=1,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))
        share_text = response.context['requisition_share_map'][str(requisition.id)]

        self.assertIn('Valor final: R$ 1.000,00', share_text)
        self.assertIn('Suborçamento 1.1', share_text)
        self.assertIn('Valor final: R$ 300,00', share_text)
        self.assertIn('Suborçamento 1.2', share_text)
        self.assertIn('Valor final: R$ 250,00', share_text)
        self.assertIn('Total orçamento principal: R$ 1.000,00', share_text)
        self.assertIn('Total suborçamentos: R$ 550,00', share_text)
        self.assertIn('Total orçamento + suborçamentos: R$ 1.550,00', share_text)

        payload = response.context['requisitions_payload'][0]
        self.assertEqual(payload['copy_budgets'][0]['group_total_display'], '1.550,00')

    def test_requisition_budget_money_inputs_use_brazilian_mask(self):
        template_path = Path(__file__).resolve().parents[1] / 'templates' / 'chamados' / 'requisicoes.html'
        content = template_path.read_text(encoding='utf-8')

        self.assertIn('class="budget-currency"', content)
        self.assertIn('Real (R$)', content)
        self.assertIn('Dolar (US$)', content)
        self.assertIn('type="text" class="budget-amount"', content)
        self.assertIn('type="text" class="budget-discount-amount"', content)
        self.assertIn('attachMoneyMask(amountInput);', content)
        self.assertIn('attachMoneyMask(discountInput);', content)

    def test_requisicoes_page_displays_budget_currency(self):
        requisition = Requisition.objects.create(
            title='Licenca em dolar',
            kind=Requisition.Kind.DIGITAL,
            request_text='Compra internacional.',
            requested_by=self.ti_user,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor externo',
            title='Licenca USD',
            currency=RequisitionBudget.Currency.USD,
            amount='99.90',
            quantity=2,
            freight_amount='10.00',
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))

        self.assertContains(response, 'Unit. US$ 99,90')
        self.assertContains(response, 'Total US$ 209,80')
        share_text = response.context['requisition_share_map'][str(requisition.id)]
        self.assertIn('Valor final: US$ 209,80', share_text)

    def test_requisicoes_page_groups_sub_budgets_under_parent_summary(self):
        requisition = Requisition.objects.create(
            title='Compra com suborcamento',
            kind=Requisition.Kind.FISICA,
            request_text='Compra com servicos adicionais.',
            requested_by=self.ti_user,
        )
        root_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='MercadoLivre',
            title='Notebook',
            amount='1799.99',
            quantity=1,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            parent_budget=root_budget,
            store_name='MercadoLivre',
            title='Memoria adicional',
            amount='159.20',
            quantity=1,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))

        self.assertContains(response, 'class="requisition-budget-group has-sub-budgets"', count=1)
        self.assertContains(response, 'requisition-budget-chip sub', count=1)
        self.assertContains(response, 'Orçamento principal')
        self.assertContains(response, 'Suborçamento')
        self.assertContains(response, 'Total R$ 1.799,99')
        self.assertContains(response, 'Total R$ 159,20')

    def test_requisicoes_page_groups_same_store_root_budgets_in_general_summary(self):
        requisition = Requisition.objects.create(
            title='Pedido tablet',
            kind=Requisition.Kind.FISICA,
            request_text='Pedido com item principal e complementos.',
            requested_by=self.ti_user,
        )
        main_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='MercadoLivre',
            title='Tablet',
            amount='1799.99',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.APROVADO,
        )
        case_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='MercadoLivre',
            title='Capa',
            amount='159.20',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )
        film_budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='MercadoLivre',
            title='Peliculas',
            amount='22.29',
            quantity=2,
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))
        main_budget.refresh_from_db()
        case_budget.refresh_from_db()
        film_budget.refresh_from_db()

        self.assertContains(response, 'class="requisition-budget-group has-sub-budgets"', count=1)
        self.assertContains(response, 'requisition-budget-chip sub', count=2)
        self.assertEqual(main_budget.approval_status, RequisitionBudget.ApprovalStatus.APROVADO)
        self.assertEqual(case_budget.approval_status, RequisitionBudget.ApprovalStatus.APROVADO)
        self.assertEqual(film_budget.approval_status, RequisitionBudget.ApprovalStatus.APROVADO)
        summary = response.context['requisition_rows'][0]['budget_summaries'][0]
        self.assertEqual(summary['approval_status'], RequisitionBudget.ApprovalStatus.APROVADO)
        self.assertEqual(
            [item['approval_status'] for item in summary['sub_summaries']],
            [RequisitionBudget.ApprovalStatus.APROVADO, RequisitionBudget.ApprovalStatus.APROVADO],
        )
        self.assertContains(response, 'Total R$ 1.799,99')
        self.assertContains(response, 'Total R$ 159,20')
        self.assertContains(response, 'Total R$ 44,58')

    def test_requisicoes_page_keeps_different_store_root_budgets_separate(self):
        requisition = Requisition.objects.create(
            title='Monitor - Planejamento',
            kind=Requisition.Kind.FISICA,
            request_text='Comparar fornecedores.',
            requested_by=self.ti_user,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='MercadoLivre',
            title='Monitor',
            amount='429.00',
            quantity=2,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Kabum',
            title='Monitor',
            amount='499.99',
            quantity=2,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))

        self.assertContains(response, 'class="requisition-budget-group "', count=2)
        self.assertNotContains(response, 'class="requisition-budget-group has-sub-budgets"')
        self.assertNotContains(response, 'requisition-budget-chip sub')

    def test_monthly_requisition_copy_uses_only_approved_budgets(self):
        april_requisition = Requisition.objects.create(
            title='Compra de bateria',
            kind=Requisition.Kind.FISICA,
            request_text='Baterias para nobreak.',
            requested_by=self.ti_user,
            requested_at=date(2026, 4, 12),
        )
        RequisitionBudget.objects.create(
            requisition=april_requisition,
            store_name='Pinha',
            title='Bateria 12V',
            amount='500.00',
            quantity=2,
            freight_amount='25.00',
            discount_amount='15.00',
            approval_status=RequisitionBudget.ApprovalStatus.APROVADO,
        )
        RequisitionBudget.objects.create(
            requisition=april_requisition,
            store_name='Gaspar',
            title='Bateria pendente',
            amount='800.00',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
        )
        may_requisition = Requisition.objects.create(
            title='Compra de memoria',
            kind=Requisition.Kind.FISICA,
            request_text='Memoria para desktop.',
            requested_by=self.ti_user,
            requested_at=date(2026, 5, 5),
        )
        RequisitionBudget.objects.create(
            requisition=may_requisition,
            store_name='Loja Maio',
            title='Memoria',
            amount='300.00',
            quantity=1,
            approval_status=RequisitionBudget.ApprovalStatus.APROVADO,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(
            reverse('chamados_requisicoes_monthly_copy'),
            data={'month': '2026-04'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['ok'])
        self.assertEqual(payload['total_display'], '1.010,00')
        self.assertEqual(payload['requisition_count'], 1)
        self.assertEqual(payload['approved_budget_count'], 1)
        self.assertEqual(payload['completed_service_count'], 0)
        self.assertEqual(payload['contract_count'], 0)
        self.assertIn('Requisições aprovadas - 04/2026', payload['text'])
        self.assertIn('REQ-', payload['text'])
        self.assertIn('Pinha', payload['text'])
        self.assertIn('Valor final: R$ 1.010,00', payload['text'])
        self.assertIn('Total geral do mês: R$ 1.010,00', payload['text'])
        self.assertNotIn('Descrição:', payload['text'])
        self.assertNotIn('Baterias para nobreak.', payload['text'])
        self.assertIn('Resumo mensal TI - 04/2026', payload['html'])
        self.assertIn('Orçamentos aprovados', payload['html'])
        self.assertIn('Total geral', payload['html'])
        self.assertIn('R$ 1.010,00', payload['html'])
        self.assertNotIn('Gaspar', payload['text'])
        self.assertNotIn('Loja Maio', payload['text'])

    def test_monthly_requisition_copy_includes_services_and_contracts(self):
        CompletedServiceEntry.objects.create(
            service_name='Manutencao nobreak',
            company='Energia Segura',
            description='Servico executado.',
            service_date=date(2026, 4, 15),
            amount='250.00',
            created_by=self.ti_user,
        )
        CompletedServiceEntry.objects.create(
            service_name='Servico fora do mes',
            company='Outra empresa',
            description='Nao deve entrar.',
            service_date=date(2026, 5, 1),
            amount='999.00',
            created_by=self.ti_user,
        )
        ContractEntry.objects.create(
            name='Contrato mensal ativo',
            notes='',
            amount='100.00',
            contract_start=date(2026, 3, 1),
            contract_end=date(2026, 5, 31),
            payment_method='Boleto',
            payment_schedule=ContractEntry.PaymentSchedule.MENSAL,
            created_by=self.ti_user,
        )
        ContractEntry.objects.create(
            name='Contrato pagamento unico',
            notes='',
            amount='300.00',
            contract_start=date(2026, 4, 10),
            contract_end=date(2026, 4, 10),
            payment_method='Pix',
            payment_schedule=ContractEntry.PaymentSchedule.PAGAMENTO_UNICO,
            created_by=self.ti_user,
        )
        ContractEntry.objects.create(
            name='Contrato anual',
            notes='',
            amount='1200.00',
            contract_start=date(2025, 4, 20),
            contract_end=date(2027, 4, 20),
            payment_method='Cartao',
            card_final='1234',
            payment_schedule=ContractEntry.PaymentSchedule.ANUAL,
            created_by=self.ti_user,
        )
        ContractEntry.objects.create(
            name='Contrato mensal fora',
            notes='',
            amount='777.00',
            contract_start=date(2026, 5, 1),
            contract_end=date(2026, 6, 1),
            payment_method='Boleto',
            payment_schedule=ContractEntry.PaymentSchedule.MENSAL,
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(
            reverse('chamados_requisicoes_monthly_copy'),
            data={'month': '2026-04'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['ok'])
        self.assertEqual(payload['completed_service_count'], 1)
        self.assertEqual(payload['contract_count'], 1)
        self.assertEqual(payload['total_display'], '550,00')
        self.assertIn('Serviços feitos no mês', payload['text'])
        self.assertIn('Manutencao nobreak', payload['text'])
        self.assertIn('Contratos do mês', payload['text'])
        self.assertIn('Contrato pagamento unico', payload['text'])
        self.assertIn('Pagamento único', payload['text'])
        self.assertNotIn('Contrato mensal ativo', payload['text'])
        self.assertNotIn('Contrato anual', payload['text'])
        self.assertNotIn('Contrato mensal fora', payload['text'])
        self.assertNotIn('Servico fora do mes', payload['text'])
        self.assertIn('Serviços feitos', payload['html'])
        self.assertIn('Contratos do mês', payload['html'])

    def test_monthly_requisition_copy_requires_valid_month(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(
            reverse('chamados_requisicoes_monthly_copy'),
            data={'month': '04/2026'},
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()['ok'])

    def test_requisition_total_uses_unit_amount_times_quantity(self):
        requisition = Requisition.objects.create(
            title='Compra de cadeiras',
            kind=Requisition.Kind.FISICA,
            request_text='Reposicao do administrativo.',
            requested_by=self.ti_user,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor principal',
            title='Cadeira presidente',
            amount='850.00',
            quantity=2,
            notes='Fornecedor principal',
        )
        root_budget = RequisitionBudget.objects.get(requisition=requisition, parent_budget__isnull=True)
        RequisitionBudget.objects.create(
            requisition=requisition,
            parent_budget=root_budget,
            title='Montagem',
            amount='120.00',
            quantity=3,
            notes='Servico adicional',
        )

        self.assertEqual(requisition.budget_total, Decimal('2060.00'))

    def test_requisition_total_includes_budget_freight_amounts(self):
        requisition = Requisition.objects.create(
            title='Compra com frete',
            kind=Requisition.Kind.FISICA,
            request_text='Entrega para filial.',
            requested_by=self.ti_user,
        )
        RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor frete',
            title='Item principal',
            amount='1000.00',
            quantity=2,
            freight_amount='150.50',
        )

        self.assertEqual(requisition.budget_total, Decimal('2150.50'))

    def test_requisition_save_accepts_brazilian_budget_freight_amount(self):
        self.client.login(username='usuario.ti', password='senha@123')
        payload = json.dumps(
            [
                {
                    'id': '',
                    'temp_key': 'tmp_root_freight',
                    'parent_ref': '',
                    'store_name': 'Fornecedor Y',
                    'title': 'Switch',
                    'amount': '1200.00',
                    'quantity': '1',
                    'freight_amount': '1.250,40',
                    'discount_amount': '0',
                    'approval_status': RequisitionBudget.ApprovalStatus.PENDENTE,
                    'receipt_status': RequisitionBudget.ReceiptStatus.PENDENTE,
                    'received_quantity': '0',
                    'notes': '',
                    'file_key': 'budget_file_tmp_root_freight',
                    'clear_file': False,
                }
            ]
        )

        response = self.client.post(
            reverse('chamados_requisicoes_save'),
            data={
                'title': 'Compra com frete brasileiro',
                'kind': Requisition.Kind.FISICA,
                'request_text': 'Teste de frete.',
                'budgets_payload': payload,
            },
        )

        self.assertRedirects(response, reverse('chamados_requisicoes'))
        requisition = Requisition.objects.get(title='Compra com frete brasileiro')
        budget = requisition.budgets.get()
        self.assertEqual(str(budget.freight_amount), '1250.40')
        self.assertEqual(requisition.budget_total, Decimal('2450.40'))

    def test_requisition_save_accepts_budget_currency_usd(self):
        self.client.login(username='usuario.ti', password='senha@123')
        payload = json.dumps(
            [
                {
                    'id': '',
                    'temp_key': 'tmp_root_usd',
                    'parent_ref': '',
                    'store_name': 'Fornecedor externo',
                    'title': 'Licenca internacional',
                    'currency': RequisitionBudget.Currency.USD,
                    'amount': '99.90',
                    'quantity': '2',
                    'freight_amount': '10.00',
                    'discount_amount': '0',
                    'approval_status': RequisitionBudget.ApprovalStatus.PENDENTE,
                    'receipt_status': RequisitionBudget.ReceiptStatus.PENDENTE,
                    'received_quantity': '0',
                    'notes': '',
                    'file_key': 'budget_file_tmp_root_usd',
                    'clear_file': False,
                }
            ]
        )

        response = self.client.post(
            reverse('chamados_requisicoes_save'),
            data={
                'title': 'Compra em dolar',
                'kind': Requisition.Kind.DIGITAL,
                'request_text': 'Licenca cotada em USD.',
                'budgets_payload': payload,
            },
        )

        self.assertRedirects(response, reverse('chamados_requisicoes'))
        budget = RequisitionBudget.objects.get(title='Licenca internacional')
        history = budget.history_entries.get()
        self.assertEqual(budget.currency, RequisitionBudget.Currency.USD)
        self.assertEqual(history.currency, RequisitionBudget.Currency.USD)
        self.assertEqual(budget.final_total, Decimal('209.80'))

    def test_sync_legacy_requisition_statuses_promotes_imported_requisition(self):
        requisition = Requisition.objects.create(
            code='LEG-REQ-00007',
            title='Requisicao importada',
            kind=Requisition.Kind.FISICA,
            request_text='[ERP-TI-REQ-ID:7]',
            status=Requisition.Status.PENDENTE_APROVACAO,
            requested_by=self.ti_user,
        )

        with TemporaryDirectory() as temp_dir:
            legacy_path = Path(temp_dir) / 'legacy.sqlite3'
            connection = sqlite3.connect(legacy_path)
            connection.execute(
                """
                CREATE TABLE core_requisition (
                    id INTEGER PRIMARY KEY,
                    status TEXT,
                    approved_at TEXT,
                    partially_received_at TEXT,
                    received_at TEXT
                )
                """
            )
            connection.execute(
                """
                INSERT INTO core_requisition (id, status, approved_at, partially_received_at, received_at)
                VALUES (7, 'approved', '2026-04-01', NULL, NULL)
                """
            )
            connection.commit()
            connection.close()

            call_command('sync_legacy_requisition_statuses', source=str(legacy_path))

        requisition.refresh_from_db()
        self.assertEqual(requisition.status, Requisition.Status.APROVADA)
        self.assertEqual(str(requisition.approved_at), '2026-04-01')
        self.assertTrue(
            RequisitionUpdate.objects.filter(
                requisition=requisition,
                status_to=Requisition.Status.APROVADA,
                message__icontains='Status sincronizado do legado ERP-TI',
            ).exists()
        )

    def test_sync_legacy_requisition_statuses_does_not_downgrade_imported_requisition(self):
        requisition = Requisition.objects.create(
            code='LEG-REQ-00009',
            title='Requisicao importada entregue',
            kind=Requisition.Kind.FISICA,
            request_text='[ERP-TI-REQ-ID:9]',
            status=Requisition.Status.ENTREGUE,
            requested_by=self.ti_user,
            approved_at=date(2026, 4, 1),
            received_at=date(2026, 4, 5),
        )

        with TemporaryDirectory() as temp_dir:
            legacy_path = Path(temp_dir) / 'legacy.sqlite3'
            connection = sqlite3.connect(legacy_path)
            connection.execute(
                """
                CREATE TABLE core_requisition (
                    id INTEGER PRIMARY KEY,
                    status TEXT,
                    approved_at TEXT,
                    partially_received_at TEXT,
                    received_at TEXT
                )
                """
            )
            connection.execute(
                """
                INSERT INTO core_requisition (id, status, approved_at, partially_received_at, received_at)
                VALUES (9, 'approved', '2026-04-01', NULL, NULL)
                """
            )
            connection.commit()
            connection.close()

            call_command('sync_legacy_requisition_statuses', source=str(legacy_path))

        requisition.refresh_from_db()
        self.assertEqual(requisition.status, Requisition.Status.ENTREGUE)
        self.assertEqual(str(requisition.received_at), '2026-04-05')
        self.assertFalse(
            RequisitionUpdate.objects.filter(
                requisition=requisition,
                message__icontains='Status sincronizado do legado ERP-TI',
            ).exists()
        )

    def test_import_erp_ti_data_imports_requisition_quote_quantity(self):
        with TemporaryDirectory() as temp_dir:
            legacy_path = Path(temp_dir) / 'legacy.sqlite3'
            connection = sqlite3.connect(legacy_path)
            connection.execute(
                """
                CREATE TABLE core_requisition (
                    id INTEGER PRIMARY KEY,
                    request TEXT,
                    quantity INTEGER,
                    unit_value DECIMAL,
                    total_value DECIMAL,
                    requested_at TEXT,
                    approved_at TEXT,
                    received_at TEXT,
                    invoice TEXT,
                    approved_by_2 TEXT,
                    req_type TEXT,
                    location TEXT,
                    link TEXT,
                    created_at TEXT,
                    updated_at TEXT,
                    status TEXT,
                    title TEXT,
                    kind TEXT,
                    delivered_quantity INTEGER,
                    partially_received_at TEXT
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE core_requisitionquote (
                    id INTEGER PRIMARY KEY,
                    name TEXT,
                    value DECIMAL,
                    photo TEXT,
                    link TEXT,
                    created_at TEXT,
                    requisition_id INTEGER,
                    freight DECIMAL,
                    is_selected BOOL,
                    quantity INTEGER,
                    payment_installments INTEGER,
                    payment_method TEXT,
                    parent_id INTEGER
                )
                """
            )
            connection.execute(
                """
                INSERT INTO core_requisition (
                    id, request, quantity, unit_value, total_value, requested_at, approved_at,
                    received_at, invoice, approved_by_2, req_type, location, link, created_at,
                    updated_at, status, title, kind, delivered_quantity, partially_received_at
                )
                VALUES (
                    77, 'Compra de mouse', 4, 35, 140, '2026-04-01', NULL,
                    NULL, '', '', 'TI', 'Matriz', '', '2026-04-01 08:00:00',
                    '2026-04-01 08:00:00', 'pending_approval', 'Mouses USB', 'physical', 0, NULL
                )
                """
            )
            connection.execute(
                """
                INSERT INTO core_requisitionquote (
                    id, name, value, photo, link, created_at, requisition_id, freight,
                    is_selected, quantity, payment_installments, payment_method, parent_id
                )
                VALUES (
                    501, 'Mouse Logitech', 33.31, '', '', '2026-04-01 08:00:00', 77, 10.36,
                    1, 4, 1, 'Pix', NULL
                )
                """
            )
            connection.commit()
            connection.close()

            call_command(
                'import_erp_ti_data',
                source=str(legacy_path),
                owner_username='usuario.ti',
            )

        budget = RequisitionBudget.objects.get(notes__contains='[ERP-TI-QUOTE-ID:501]')
        self.assertEqual(budget.quantity, 4)
        self.assertEqual(budget.approval_status, RequisitionBudget.ApprovalStatus.APROVADO)
        self.assertEqual(str(budget.amount), '33.31')
        self.assertEqual(str(budget.freight_amount), '10.36')

    def test_sync_legacy_requisition_quantities_updates_imported_budget(self):
        requisition = Requisition.objects.create(
            code='LEG-REQ-00015',
            title='Legado quantidade',
            kind=Requisition.Kind.FISICA,
            request_text='Quantidade veio incorreta.',
            requested_by=self.ti_user,
        )
        budget = RequisitionBudget.objects.create(
            requisition=requisition,
            title='Windows Server CAL',
            amount='325.00',
            quantity=1,
            notes='Quantidade: 90\n[ERP-TI-QUOTE-ID:900]',
        )

        with TemporaryDirectory() as temp_dir:
            legacy_path = Path(temp_dir) / 'legacy.sqlite3'
            connection = sqlite3.connect(legacy_path)
            connection.execute(
                """
                CREATE TABLE core_requisitionquote (
                    id INTEGER PRIMARY KEY,
                    quantity INTEGER
                )
                """
            )
            connection.execute('INSERT INTO core_requisitionquote (id, quantity) VALUES (900, 90)')
            connection.commit()
            connection.close()

            call_command('sync_legacy_requisition_quantities', source=str(legacy_path))

        budget.refresh_from_db()
        self.assertEqual(budget.quantity, 90)

    def test_sync_legacy_requisition_budget_approvals_marks_selected_quote(self):
        requisition = Requisition.objects.create(
            code='LEG-REQ-00016',
            title='Legado aprovado',
            kind=Requisition.Kind.FISICA,
            request_text='Orcamento selecionado no legado.',
            status=Requisition.Status.PENDENTE_APROVACAO,
            requested_by=self.ti_user,
        )
        budget = RequisitionBudget.objects.create(
            requisition=requisition,
            title='Tablet Xiaomi',
            amount='1429.00',
            quantity=9,
            approval_status=RequisitionBudget.ApprovalStatus.PENDENTE,
            notes='Selecionado legado: True\n[ERP-TI-QUOTE-ID:777]',
        )

        with TemporaryDirectory() as temp_dir:
            legacy_path = Path(temp_dir) / 'legacy.sqlite3'
            connection = sqlite3.connect(legacy_path)
            connection.execute(
                """
                CREATE TABLE core_requisitionquote (
                    id INTEGER PRIMARY KEY,
                    is_selected BOOL
                )
                """
            )
            connection.execute('INSERT INTO core_requisitionquote (id, is_selected) VALUES (777, 1)')
            connection.commit()
            connection.close()

            call_command('sync_legacy_requisition_budget_approvals', source=str(legacy_path))

        budget.refresh_from_db()
        requisition.refresh_from_db()
        self.assertEqual(budget.approval_status, RequisitionBudget.ApprovalStatus.APROVADO)
        self.assertEqual(requisition.status, Requisition.Status.APROVADA)
        self.assertTrue(
            RequisitionUpdate.objects.filter(
                requisition=requisition,
                message__icontains='Orcamento aprovado sincronizado do legado ERP-TI',
            ).exists()
        )

    def test_requisition_budget_history_is_visible_in_payload(self):
        requisition = Requisition.objects.create(
            title='Compra de monitor',
            kind=Requisition.Kind.FISICA,
            request_text='Expansao de equipe.',
            requested_by=self.ti_user,
        )
        budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor E',
            title='Monitor 27',
            amount='1200.00',
            quantity=2,
            discount_amount='150.00',
            approval_status=RequisitionBudget.ApprovalStatus.APROVADO,
            receipt_status=RequisitionBudget.ReceiptStatus.PARCIAL,
            received_quantity=1,
            notes='Fornecedor E',
        )
        RequisitionBudgetHistory.objects.create(
            budget=budget,
            author=self.ti_user,
            message='Orcamento atualizado (valores).',
            store_name='Fornecedor E',
            amount='1200.00',
            quantity=2,
            line_total='2400.00',
            discount_amount='150.00',
            final_total='2250.00',
            approval_status=RequisitionBudget.ApprovalStatus.APROVADO,
            receipt_status=RequisitionBudget.ReceiptStatus.PARCIAL,
            received_quantity=1,
            remaining_quantity=1,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))
        self.assertContains(response, 'Historico')
        self.assertContains(response, 'Recebido parcial')

    def test_requisicoes_page_shows_image_thumbnail_for_budget_attachment(self):
        requisition = Requisition.objects.create(
            title='Compra de webcam',
            kind=Requisition.Kind.FISICA,
            request_text='Item para sala de reunioes.',
            requested_by=self.ti_user,
        )
        budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor D',
            title='Orcamento principal',
            amount='450.00',
            notes='Fornecedor D',
        )
        budget.evidence_file.save('print_orcamento.png', ContentFile(b'fake-image-bytes'), save=True)

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))
        self.assertContains(response, '/media/requisitions/budgets/')
        self.assertContains(response, 'budget-thumb')

    def test_requisition_save_accepts_multiple_budget_documents(self):
        self.client.login(username='usuario.ti', password='senha@123')
        payload = json.dumps(
            [
                {
                    'id': '',
                    'temp_key': 'tmp_docs',
                    'parent_ref': '',
                    'store_name': 'Fornecedor Docs',
                    'title': 'Orcamento com documentos',
                    'amount': '750.00',
                    'quantity': '1',
                    'freight_amount': '0',
                    'discount_amount': '0',
                    'approval_status': RequisitionBudget.ApprovalStatus.PENDENTE,
                    'receipt_status': RequisitionBudget.ReceiptStatus.PENDENTE,
                    'received_quantity': '0',
                    'notes': '',
                    'file_key': 'budget_file_tmp_docs',
                    'attachment_key': 'budget_attachments_tmp_docs',
                    'clear_file': False,
                },
            ]
        )

        response = self.client.post(
            reverse('chamados_requisicoes_save'),
            data={
                'title': 'Compra com documentos por orcamento',
                'kind': Requisition.Kind.FISICA,
                'request_text': 'Documentos adicionais do fornecedor.',
                'budgets_payload': payload,
                'budget_attachments_tmp_docs': [
                    SimpleUploadedFile('proposta.pdf', b'pdf-bytes', content_type='application/pdf'),
                    SimpleUploadedFile('condicoes.docx', b'docx-bytes', content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'),
                ],
            },
        )

        self.assertRedirects(response, reverse('chamados_requisicoes'))
        budget = RequisitionBudget.objects.get(title='Orcamento com documentos')
        attachments = list(budget.attachments.order_by('id'))
        self.assertEqual(len(attachments), 2)
        self.assertTrue(attachments[0].file.name.endswith('.pdf'))
        self.assertTrue(attachments[1].file.name.endswith('.docx'))

    def test_requisicoes_page_lists_budget_document_attachments(self):
        requisition = Requisition.objects.create(
            title='Compra com anexo documental',
            kind=Requisition.Kind.FISICA,
            request_text='Documentos extras do orcamento.',
            requested_by=self.ti_user,
        )
        budget = RequisitionBudget.objects.create(
            requisition=requisition,
            store_name='Fornecedor E',
            title='Orcamento documental',
            amount='520.00',
            notes='Fornecedor E',
        )
        attachment = RequisitionBudgetAttachment.objects.create(budget=budget)
        attachment.file.save('proposta_fornecedor.pdf', ContentFile(b'pdf-bytes'), save=True)

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_requisicoes'))
        self.assertContains(response, 'Documentos adicionais')
        self.assertContains(response, 'proposta_fornecedor.pdf')
        self.assertContains(response, '/media/requisitions/budget_documents/')

    def test_only_ti_can_access_insumos_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_insumos'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_insumos'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Insumos TI')

    def test_ti_can_create_and_update_insumo_record(self):
        self.client.login(username='usuario.ti', password='senha@123')
        create_response = self.client.post(
            reverse('chamados_insumos'),
            data={
                'mode': 'create',
                'item': 'Mouse',
                'date': '2026-04-10',
                'quantity': '2,00',
                'name': 'Fabiano',
                'department': 'TI',
            },
        )
        self.assertRedirects(create_response, reverse('chamados_insumos'))
        insumo = Insumo.objects.get()
        self.assertEqual(insumo.item, 'Mouse')
        self.assertEqual(str(insumo.quantity), '2.00')

        update_response = self.client.post(
            reverse('chamados_insumos'),
            data={
                'mode': 'update',
                'insumo_id': insumo.id,
                'item': 'Mouse sem fio',
                'date': '2026-04-11',
                'quantity': '3,00',
                'name': 'Fabiano',
                'department': 'TI',
            },
        )
        self.assertRedirects(update_response, reverse('chamados_insumos'))
        insumo.refresh_from_db()
        self.assertEqual(insumo.item, 'Mouse sem fio')
        self.assertEqual(str(insumo.quantity), '3.00')

    def test_ti_can_register_stock_and_output(self):
        self.client.login(username='usuario.ti', password='senha@123')
        self.client.post(
            reverse('chamados_insumos'),
            data={
                'mode': 'stock_create',
                'stock_item': 'Bateria',
                'stock_quantity': '5,00',
            },
        )
        self.client.post(
            reverse('chamados_insumos'),
            data={
                'mode': 'stock_adjust',
                'stock_item': 'Bateria',
                'stock_direction': 'dec',
                'stock_quantity': '2,00',
                'stock_target': 'Setor PCP',
                'stock_reason': 'Reposicao',
            },
        )
        response = self.client.get(reverse('chamados_insumos'))
        self.assertContains(response, 'Bateria')
        self.assertContains(response, '>3<', html=False)

    def test_only_ti_can_access_starlinks_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_starlinks'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_starlinks'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Starlinks')

    def test_ti_can_create_starlink(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_starlinks'),
            data={
                'name': 'Starlink Matriz',
                'location': 'Recepcao',
                'email': 'starlink@sidertec.com.br',
                'is_active': 'on',
                'payment_method': 'cartao',
                'card_final': '1234',
            },
        )
        self.assertRedirects(response, reverse('chamados_starlinks'))
        starlink = Starlink.objects.get()
        self.assertEqual(starlink.name, 'Starlink Matriz')
        self.assertEqual(starlink.location, 'Recepcao')
        self.assertEqual(starlink.email, 'starlink@sidertec.com.br')
        self.assertTrue(starlink.is_active)
        self.assertEqual(starlink.payment_method, Starlink.PaymentMethod.CARTAO)
        self.assertEqual(starlink.card_final, '1234')
        self.assertEqual(starlink.created_by, self.ti_user)
        self.assertEqual(starlink.password_encrypted, '')

    def test_ti_can_create_starlink_with_pix_without_card_final(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_starlinks'),
            data={
                'name': 'Starlink Filial',
                'location': 'Filial',
                'email': 'pix@sidertec.com.br',
                'is_active': 'on',
                'payment_method': 'pix',
                'card_final': '',
            },
        )
        self.assertRedirects(response, reverse('chamados_starlinks'))
        starlink = Starlink.objects.get(name='Starlink Filial')
        self.assertEqual(starlink.payment_method, Starlink.PaymentMethod.PIX)
        self.assertEqual(starlink.card_final, '')

    def test_ti_can_view_starlink_detail_without_password(self):
        starlink = Starlink.objects.create(
            name='Starlink Detalhe',
            location='PCP',
            email='detalhe@sidertec.com.br',
            is_active=True,
            payment_method=Starlink.PaymentMethod.CARTAO,
            card_final='9876',
            created_by=self.ti_user,
            password_encrypted='',
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_starlinks_detail', args=[starlink.id]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Editar dados')
        self.assertContains(response, 'Apagar')
        self.assertNotContains(response, 'Senha')

    def test_ti_can_update_starlink(self):
        starlink = Starlink.objects.create(
            name='Starlink Antiga',
            location='Almox',
            email='antiga@sidertec.com.br',
            is_active=True,
            payment_method=Starlink.PaymentMethod.CARTAO,
            card_final='1111',
            created_by=self.ti_user,
            password_encrypted='',
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_starlinks_update', args=[starlink.id]),
            data={
                'name': 'Starlink Nova',
                'location': 'Expedicao',
                'email': 'nova@sidertec.com.br',
                'payment_method': 'pix',
                'card_final': '',
                'is_active': '',
            },
        )

        self.assertRedirects(response, reverse('chamados_starlinks_detail', args=[starlink.id]))
        starlink.refresh_from_db()
        self.assertEqual(starlink.name, 'Starlink Nova')
        self.assertEqual(starlink.location, 'Expedicao')
        self.assertEqual(starlink.email, 'nova@sidertec.com.br')
        self.assertFalse(starlink.is_active)
        self.assertEqual(starlink.payment_method, Starlink.PaymentMethod.PIX)
        self.assertEqual(starlink.card_final, '')
        self.assertEqual(starlink.password_encrypted, '')

    def test_ti_can_delete_starlink(self):
        starlink = Starlink.objects.create(
            name='Starlink Apagar',
            location='Recepcao',
            email='apagar@sidertec.com.br',
            is_active=True,
            payment_method=Starlink.PaymentMethod.CARTAO,
            card_final='2222',
            created_by=self.ti_user,
            password_encrypted='',
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(reverse('chamados_starlinks_delete', args=[starlink.id]))

        self.assertRedirects(response, reverse('chamados_starlinks'))
        self.assertFalse(Starlink.objects.filter(id=starlink.id).exists())

    def test_only_ti_can_access_documentos_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_documentos'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_documentos'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Documentos')

    def test_ti_can_create_documento(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_documentos'),
            data={
                'name': 'Manual da impressora fiscal',
                'notes': 'Arquivo e observacoes para reinstalacao rapida.',
            },
        )

        self.assertRedirects(response, reverse('chamados_documentos'))
        documento = DocumentEntry.objects.get()
        self.assertEqual(documento.name, 'Manual da impressora fiscal')
        self.assertEqual(documento.notes, 'Arquivo e observacoes para reinstalacao rapida.')
        self.assertEqual(documento.created_by, self.ti_user)

    def test_ti_can_create_documento_with_attachment(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_documentos'),
            data={
                'name': 'Procedimento VPN',
                'notes': 'Passo a passo de configuracao.',
                'attachment': ContentFile(b'pdf-teste', name='procedimento_vpn.pdf'),
            },
        )

        self.assertRedirects(response, reverse('chamados_documentos'))
        documento = DocumentEntry.objects.get(name='Procedimento VPN')
        self.assertIn('procedimento_vpn', documento.attachment.name)
        self.assertTrue(documento.attachment.name.endswith('.pdf'))

    def test_only_ti_can_access_emprestimos_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_emprestimos'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_emprestimos'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Emprestimos')
        self.assertContains(response, 'Novo emprestimo')

    def test_ti_can_create_equipment_loan_and_download_term(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_emprestimos'),
            data={
                'collaborator_name': 'Alexandre Graciano',
                'collaborator_company': 'Parceiro externo',
                'collaborator_document': '123.456.789-00',
                'collaborator_email': 'alexandre@example.com',
                'collaborator_phone': '(11) 99999-9999',
                'equipment_type': 'Notebook',
                'equipment_brand': 'Dell',
                'equipment_model': 'Latitude',
                'equipment_serial': 'SN123',
                'patrimony_tag': 'TI-001',
                'accessories': 'Fonte\nMochila',
                'loan_date': '2026-05-12',
                'expected_return_date': '2026-06-12',
                'notes': 'Emprestimo para projeto externo.',
            },
        )

        self.assertRedirects(response, reverse('chamados_emprestimos'))
        loan = EquipmentLoan.objects.get()
        self.assertEqual(loan.created_by, self.ti_user)
        self.assertFalse(loan.documentation_ok)

        response = self.client.get(reverse('chamados_emprestimos_termo', args=[loan.id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('termo_emprestimo_Alexandre_Graciano', response['Content-Disposition'])
        self.assertIn('.pdf', response['Content-Disposition'])
        self.assertTrue(response.content.startswith(b'%PDF-1.4'))
        self.assertIn(b'Alexandre Graciano', response.content)
        self.assertIn(b'Notebook', response.content)

        response = self.client.get(reverse('chamados_emprestimos_termo_devolucao', args=[loan.id]))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response['Content-Type'], 'application/pdf')
        self.assertIn('termo_devolucao_Alexandre_Graciano', response['Content-Disposition'])
        self.assertTrue(response.content.startswith(b'%PDF-1.4'))

    def test_ti_can_mark_equipment_loan_as_returned(self):
        loan = EquipmentLoan.objects.create(
            collaborator_name='Alexandre Graciano',
            collaborator_company='Parceiro externo',
            equipment_type='Notebook',
            loan_date=date(2026, 5, 12),
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_emprestimos'),
            data={
                'mode': 'mark_returned',
                'loan_id': loan.id,
            },
        )

        self.assertRedirects(response, reverse('chamados_emprestimos'))
        loan.refresh_from_db()
        self.assertTrue(loan.returned)
        self.assertIsNotNone(loan.returned_at)
        self.assertEqual(loan.returned_by, self.ti_user)

    def test_ti_can_upload_signed_equipment_loan_document_and_mark_ok(self):
        loan = EquipmentLoan.objects.create(
            collaborator_name='Alexandre Graciano',
            collaborator_company='Parceiro externo',
            equipment_type='Notebook',
            loan_date=date(2026, 5, 12),
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_emprestimos'),
            data={
                'mode': 'upload_signed',
                'loan_id': loan.id,
                'signed_document': SimpleUploadedFile('termo_assinado.pdf', b'pdf-assinado', content_type='application/pdf'),
            },
        )

        self.assertRedirects(response, reverse('chamados_emprestimos'))
        loan.refresh_from_db()
        self.assertTrue(loan.documentation_ok)
        self.assertIsNotNone(loan.documentation_ok_at)
        self.assertTrue(loan.signed_document.name.endswith('.pdf'))

    def test_ti_can_upload_attendant_signature_for_equipment_loan_pdf(self):
        loan = EquipmentLoan.objects.create(
            collaborator_name='Alexandre Graciano',
            collaborator_company='Parceiro externo',
            equipment_type='Notebook',
            loan_date=date(2026, 5, 12),
            created_by=self.ti_user,
        )
        signature_png = Path('Logo Verde.png').read_bytes()

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_emprestimos'),
            data={
                'mode': 'upload_attendant_signature',
                'loan_id': loan.id,
                'attendant_signature': SimpleUploadedFile('assinatura.png', signature_png, content_type='image/png'),
            },
        )

        self.assertRedirects(response, reverse('chamados_emprestimos'))
        loan.refresh_from_db()
        self.assertTrue(loan.attendant_signature.name.endswith('.png'))

        response = self.client.get(reverse('chamados_emprestimos_termo', args=[loan.id]))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.content.startswith(b'%PDF-1.4'))

    def test_ti_can_attach_equipment_loan_photos_on_create_and_later(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_emprestimos'),
            data={
                'collaborator_name': 'Alexandre Graciano',
                'collaborator_company': 'Parceiro externo',
                'equipment_type': 'Notebook',
                'loan_date': '2026-05-12',
                'photos': [
                    SimpleUploadedFile('frente.jpg', b'fake-jpg-1', content_type='image/jpeg'),
                    SimpleUploadedFile('verso.png', b'fake-png-1', content_type='image/png'),
                ],
            },
        )

        self.assertRedirects(response, reverse('chamados_emprestimos'))
        loan = EquipmentLoan.objects.get()
        self.assertEqual(loan.photos.count(), 2)

        response = self.client.post(
            reverse('chamados_emprestimos'),
            data={
                'mode': 'add_photos',
                'loan_id': loan.id,
                'photos': [
                    SimpleUploadedFile('serie.jpg', b'fake-jpg-2', content_type='image/jpeg'),
                ],
            },
        )

        self.assertRedirects(response, reverse('chamados_emprestimos'))
        self.assertEqual(EquipmentLoanPhoto.objects.filter(loan=loan).count(), 3)
        response = self.client.get(reverse('chamados_emprestimos'))
        self.assertContains(response, 'equipment-loan-photo-list')
        self.assertContains(response, 'Anexar fotos')

    def _workspace_csv_upload(self, content: str):
        return SimpleUploadedFile(
            'workspace.csv',
            content.encode('utf-8'),
            content_type='text/csv',
        )

    def test_only_ti_can_access_emails_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_emails'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_emails'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Emails')
        self.assertContains(response, 'Importar / atualizar CSV')
        self.assertContains(response, 'Copiar emails')
        self.assertContains(response, 'data-sort-key="email"', html=False)
        self.assertContains(response, 'buildEmailsClipboardText', html=False)

    def test_only_ti_can_access_ramais_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_ramais'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_ramais'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Ramais')
        self.assertContains(response, 'Novo ramal')
        self.assertContains(response, 'phoneExtensionSearchInput')
        self.assertContains(response, 'phoneExtensionDepartmentFilter')
        self.assertContains(response, 'phoneExtensionExtensionFilter')
        self.assertContains(response, 'phoneExtensionEmailFilter')

    def test_ramais_page_displays_filter_options(self):
        PhoneExtension.objects.create(
            department='TI',
            name='Fabiano Polone',
            phone='(16) 3353-8390',
            extension='8390',
            email='fabiano.polone@sidertec.com.br',
            created_by=self.ti_user,
        )
        PhoneExtension.objects.create(
            department='Financeiro',
            name='Gabriel Cordeiro',
            phone='(16) 3353-8430',
            extension='8430',
            email='',
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_ramais'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<option value="TI">TI</option>', html=True)
        self.assertContains(response, '<option value="Financeiro">Financeiro</option>', html=True)
        self.assertContains(response, '<option value="8390">8390</option>', html=True)
        self.assertContains(response, '<option value="8430">8430</option>', html=True)
        self.assertContains(response, 'data-has-email="yes"', html=False)
        self.assertContains(response, 'data-has-email="no"', html=False)

    def test_ti_can_create_phone_extension(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_ramais'),
            data={
                'department': 'Recepcao',
                'name': 'Recepcao Matriz',
                'phone': '(16) 3333-0000',
                'extension': '201',
                'email': 'recepcao@sidertec.com.br',
            },
        )

        self.assertRedirects(response, reverse('chamados_ramais'))
        extension = PhoneExtension.objects.get()
        self.assertEqual(extension.name, 'Recepcao Matriz')
        self.assertEqual(extension.department, 'Recepcao')
        self.assertEqual(extension.phone, '(16) 3333-0000')
        self.assertEqual(extension.extension, '201')
        self.assertEqual(extension.email, 'recepcao@sidertec.com.br')
        self.assertEqual(extension.created_by, self.ti_user)

        response = self.client.get(reverse('chamados_ramais'))
        self.assertContains(response, 'Recepcao Matriz')
        self.assertContains(response, '201')
        self.assertContains(response, 'recepcao@sidertec.com.br')

    def test_ti_can_update_phone_extension(self):
        extension = PhoneExtension.objects.create(
            department='Recepcao',
            name='Recepcao Matriz',
            phone='(16) 3333-0000',
            extension='201',
            email='recepcao@sidertec.com.br',
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_ramais'),
            data={
                'mode': 'update',
                'extension_id': extension.id,
                'department': 'TI',
                'name': 'Recepcao Atualizada',
                'phone': '(16) 3333-1111',
                'extension': '202',
                'email': 'recepcao@sidertec.com.br\nportaria@sidertec.com.br',
            },
        )

        self.assertRedirects(response, reverse('chamados_ramais'))
        extension.refresh_from_db()
        self.assertEqual(extension.department, 'TI')
        self.assertEqual(extension.name, 'Recepcao Atualizada')
        self.assertEqual(extension.phone, '(16) 3333-1111')
        self.assertEqual(extension.extension, '202')
        self.assertEqual(extension.email, 'recepcao@sidertec.com.br\nportaria@sidertec.com.br')

    def test_ti_can_delete_phone_extension(self):
        extension = PhoneExtension.objects.create(
            department='Recepcao',
            name='Recepcao Matriz',
            phone='(16) 3333-0000',
            extension='201',
            email='recepcao@sidertec.com.br',
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_ramais'),
            data={
                'mode': 'delete',
                'extension_id': extension.id,
            },
        )

        self.assertRedirects(response, reverse('chamados_ramais'))
        self.assertFalse(PhoneExtension.objects.filter(id=extension.id).exists())

    def test_import_ramais_command_imports_xlsx(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Departamento', 'Colaborador', 'Telefone', 'Ramal', 'Email'])
        sheet.append(['Almoxarifado', 'Manoel Santos', '(16) 3353-8423', '8423', 'manoel.santos@sidertec.com.br'])
        sheet.append(['TI', 'Fabiano Polone', '(16) 3353-8390', '8390', 'fabiano.polone@sidertec.com.br'])

        with TemporaryDirectory() as tmpdir:
            source = Path(tmpdir) / 'ramais.xlsx'
            workbook.save(source)
            call_command('import_ramais', '--source', str(source), '--created-by', self.ti_user.username)

        self.assertEqual(PhoneExtension.objects.count(), 2)
        ramal = PhoneExtension.objects.get(email='manoel.santos@sidertec.com.br')
        self.assertEqual(ramal.department, 'Almoxarifado')
        self.assertEqual(ramal.name, 'Manoel Santos')
        self.assertEqual(ramal.phone, '(16) 3353-8423')
        self.assertEqual(ramal.extension, '8423')
        self.assertEqual(ramal.created_by, self.ti_user)

    def test_ti_can_import_google_workspace_email_csv(self):
        csv_content = (
            'First Name [Required],Last Name [Required],Email Address [Required],Password [Required],'
            'Status [READ ONLY],Last Sign In [READ ONLY],Email Usage [READ ONLY],Drive Usage [READ ONLY],'
            'Storage Used [READ ONLY],Licenses [READ ONLY]\n'
            'Fabiano,Polone,fabiano.polone@sidertec.com.br,****,Active,2026/04/28 08:30:00,'
            '1.20GB,0.50GB,1.70GB,1010020029\n'
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_emails'),
            data={'csv_file': self._workspace_csv_upload(csv_content)},
        )

        self.assertRedirects(response, reverse('chamados_emails'))
        account = GoogleWorkspaceEmail.objects.get(email='fabiano.polone@sidertec.com.br')
        self.assertEqual(account.first_name, 'Fabiano')
        self.assertEqual(account.last_name, 'Polone')
        self.assertEqual(account.status, 'Active')
        self.assertEqual(account.last_sign_in, '2026/04/28 08:30:00')
        self.assertEqual(account.email_usage, '1.20GB')
        self.assertEqual(account.drive_usage, '0.50GB')
        self.assertEqual(account.storage_used, '1.70GB')
        self.assertEqual(account.license_code, '1010020029')
        self.assertEqual(account.imported_by, self.ti_user)

    def test_google_workspace_email_import_updates_existing_records(self):
        GoogleWorkspaceEmail.objects.create(
            first_name='Fabiano',
            last_name='Polone',
            email='fabiano.polone@sidertec.com.br',
            status='Active',
            last_sign_in='2026/04/20 08:30:00',
            email_usage='1.20GB',
            drive_usage='0.50GB',
            storage_used='1.70GB',
            license_code='1010020029',
            imported_by=self.ti_user,
        )
        csv_content = (
            'First Name [Required],Last Name [Required],Email Address [Required],Password [Required],'
            'Status [READ ONLY],Last Sign In [READ ONLY],Email Usage [READ ONLY],Drive Usage [READ ONLY],'
            'Storage Used [READ ONLY],Licenses [READ ONLY]\n'
            'Fabiano,Polone,fabiano.polone@sidertec.com.br,****,Suspended,2026/04/29 10:00:00,'
            '2.00GB,0.70GB,2.70GB,1010020029\n'
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_emails'),
            data={'csv_file': self._workspace_csv_upload(csv_content)},
            follow=True,
        )

        self.assertContains(response, '0 criados, 1 atualizados, 0 sem alteracao')
        self.assertEqual(GoogleWorkspaceEmail.objects.count(), 1)
        account = GoogleWorkspaceEmail.objects.get(email='fabiano.polone@sidertec.com.br')
        self.assertEqual(account.status, 'Suspended')
        self.assertEqual(account.last_sign_in, '2026/04/29 10:00:00')
        self.assertEqual(account.storage_used, '2.70GB')

    def test_google_workspace_email_import_removes_records_missing_from_csv(self):
        GoogleWorkspaceEmail.objects.create(
            first_name='Fabiano',
            last_name='Polone',
            email='fabiano.polone@sidertec.com.br',
            status='Active',
            imported_by=self.ti_user,
        )
        GoogleWorkspaceEmail.objects.create(
            first_name='Antigo',
            last_name='Usuario',
            email='antigo.usuario@sidertec.com.br',
            status='Active',
            imported_by=self.ti_user,
        )
        csv_content = (
            'First Name [Required],Last Name [Required],Email Address [Required],Password [Required],'
            'Status [READ ONLY],Last Sign In [READ ONLY],Email Usage [READ ONLY],Drive Usage [READ ONLY],'
            'Storage Used [READ ONLY],Licenses [READ ONLY]\n'
            'Fabiano,Polone,fabiano.polone@sidertec.com.br,****,Active,2026/04/29 10:00:00,'
            '2.00GB,0.70GB,2.70GB,1010020029\n'
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_emails'),
            data={'csv_file': self._workspace_csv_upload(csv_content)},
            follow=True,
        )

        self.assertContains(response, '1 removidos')
        self.assertTrue(GoogleWorkspaceEmail.objects.filter(email='fabiano.polone@sidertec.com.br').exists())
        self.assertFalse(GoogleWorkspaceEmail.objects.filter(email='antigo.usuario@sidertec.com.br').exists())

    def test_google_workspace_email_import_empty_file_keeps_current_records(self):
        GoogleWorkspaceEmail.objects.create(
            first_name='Fabiano',
            last_name='Polone',
            email='fabiano.polone@sidertec.com.br',
            status='Active',
            imported_by=self.ti_user,
        )
        csv_content = (
            'First Name [Required],Last Name [Required],Email Address [Required],Password [Required],'
            'Status [READ ONLY],Last Sign In [READ ONLY],Email Usage [READ ONLY],Drive Usage [READ ONLY],'
            'Storage Used [READ ONLY],Licenses [READ ONLY]\n'
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_emails'),
            data={'csv_file': self._workspace_csv_upload(csv_content)},
            follow=True,
        )

        self.assertContains(response, 'Nenhum email valido foi encontrado no CSV')
        self.assertEqual(GoogleWorkspaceEmail.objects.count(), 1)

    def test_emails_page_searches_workspace_accounts(self):
        GoogleWorkspaceEmail.objects.create(
            first_name='Fabiano',
            last_name='Polone',
            email='fabiano.polone@sidertec.com.br',
            status='Active',
            last_sign_in='2026/04/29 10:00:00',
            email_usage='2.00GB',
            drive_usage='0.70GB',
            storage_used='2.70GB',
            license_code='1010020029',
            imported_by=self.ti_user,
        )
        GoogleWorkspaceEmail.objects.create(
            first_name='Albeni',
            last_name='Silva',
            email='albeni.silva@sidertec.com.br',
            status='Active',
            last_sign_in='2026/04/28 01:44:42',
            email_usage='0.94GB',
            drive_usage='0.09GB',
            storage_used='1.03GB',
            license_code='1010020029',
            imported_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_emails') + '?q=fabiano active')

        self.assertContains(response, 'fabiano.polone@sidertec.com.br')
        self.assertNotContains(response, 'albeni.silva@sidertec.com.br')

    def test_only_ti_can_access_servicos_feitos_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_servicos_feitos'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_servicos_feitos'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Servicos feitos')

    def test_ti_can_create_servico_feito_with_attachment_and_brazilian_amount(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_servicos_feitos'),
            data={
                'service_name': 'Manutencao nobreak',
                'company': 'Energia Segura Ltda',
                'description': 'Troca de baterias e teste de autonomia.',
                'service_date': '2026-04-15',
                'attachments': ContentFile(b'ordem-servico', name='os_nobreak.pdf'),
                'amount': '1.250,40',
            },
        )

        self.assertRedirects(response, reverse('chamados_servicos_feitos'))
        entry = CompletedServiceEntry.objects.get()
        self.assertEqual(entry.service_name, 'Manutencao nobreak')
        self.assertEqual(entry.company, 'Energia Segura Ltda')
        self.assertEqual(entry.description, 'Troca de baterias e teste de autonomia.')
        self.assertEqual(entry.service_date, date(2026, 4, 15))
        self.assertEqual(str(entry.amount), '1250.40')
        attachment = entry.attachments.get()
        self.assertTrue(attachment.file.name.endswith('.pdf'))
        self.assertEqual(entry.created_by, self.ti_user)

    def test_ti_can_create_servico_feito_with_multiple_attachments(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_servicos_feitos'),
            data={
                'service_name': 'Instalacao cameras',
                'company': 'Seguranca Total',
                'description': 'Instalacao e validacao.',
                'service_date': '2026-04-16',
                'attachments': [
                    ContentFile(b'nota-fiscal', name='nota.pdf'),
                    ContentFile(b'fotos-servico', name='fotos.zip'),
                ],
                'amount': '850,00',
            },
        )

        self.assertRedirects(response, reverse('chamados_servicos_feitos'))
        entry = CompletedServiceEntry.objects.get(service_name='Instalacao cameras')
        attachments = list(entry.attachments.order_by('id'))
        self.assertEqual(len(attachments), 2)
        self.assertTrue(attachments[0].file.name.endswith('.pdf'))
        self.assertTrue(attachments[1].file.name.endswith('.zip'))

    def test_ti_can_update_servico_feito_service_date(self):
        entry = CompletedServiceEntry.objects.create(
            service_name='Troca de bateria',
            company='Energia Segura',
            description='Troca concluida.',
            service_date=date(2026, 4, 10),
            amount='300.00',
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_servicos_feitos'),
            data={
                'mode': 'update_service_date',
                'entry_id': entry.id,
                'service_date': '2026-04-20',
            },
        )

        self.assertRedirects(response, reverse('chamados_servicos_feitos'))
        entry.refresh_from_db()
        self.assertEqual(entry.service_date, date(2026, 4, 20))

    def test_ti_can_add_attachments_to_existing_servico_feito(self):
        entry = CompletedServiceEntry.objects.create(
            service_name='Revisao firewall',
            company='Seguranca Redes',
            description='Revisao de regras concluida.',
            service_date=date(2026, 4, 21),
            amount='1200.00',
            created_by=self.ti_user,
        )
        CompletedServiceAttachment.objects.create(
            service=entry,
            file=ContentFile(b'evidencia-antiga', name='antes.pdf'),
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_servicos_feitos'),
            data={
                'mode': 'add_attachments',
                'entry_id': entry.id,
                'attachments': [
                    ContentFile(b'evidencia-nova', name='depois.pdf'),
                    ContentFile(b'fotos', name='fotos.zip'),
                ],
            },
        )

        self.assertRedirects(response, reverse('chamados_servicos_feitos'))
        attachments = list(entry.attachments.order_by('id'))
        self.assertEqual(len(attachments), 3)
        self.assertTrue(attachments[0].file.name.endswith('.pdf'))
        self.assertTrue(attachments[1].file.name.endswith('.pdf'))
        self.assertTrue(attachments[2].file.name.endswith('.zip'))

    def test_servicos_feitos_page_displays_amount_in_brazilian_format(self):
        CompletedServiceEntry.objects.create(
            service_name='Cabeamento rack',
            company='Infra Redes',
            description='Organizacao e identificacao.',
            service_date=date(2026, 4, 17),
            amount='2499.90',
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_servicos_feitos'))

        self.assertContains(response, 'R$ 2.499,90')
        self.assertContains(response, '2026-04-17')

    def test_servicos_feitos_page_lists_multiple_attachments(self):
        entry = CompletedServiceEntry.objects.create(
            service_name='Backup servidor',
            company='Infra Redes',
            description='Backup completo.',
            service_date=date(2026, 4, 18),
            amount='500.00',
            created_by=self.ti_user,
        )
        CompletedServiceAttachment.objects.create(
            service=entry,
            file=ContentFile(b'relatorio', name='relatorio.pdf'),
        )
        CompletedServiceAttachment.objects.create(
            service=entry,
            file=ContentFile(b'evidencia', name='evidencia.png'),
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_servicos_feitos'))

        self.assertContains(response, 'Abrir 1')
        self.assertContains(response, 'Abrir 2')

    def test_only_ti_can_access_contratos_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_contratos'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_contratos'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Contratos')

    def test_ti_can_create_contrato(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_contratos'),
            data={
                'name': 'Contrato Microsoft 365',
                'notes': 'Renovacao anual do licenciamento corporativo.',
                'amount': '2499.90',
                'contract_start': '2026-01-01',
                'contract_end': '2026-12-31',
                'payment_method': 'Cartao',
                'card_final': '1234',
                'payment_schedule': 'mensal',
            },
        )

        self.assertRedirects(response, reverse('chamados_contratos'))
        contrato = ContractEntry.objects.get()
        self.assertEqual(contrato.name, 'Contrato Microsoft 365')
        self.assertEqual(contrato.notes, 'Renovacao anual do licenciamento corporativo.')
        self.assertEqual(str(contrato.amount), '2499.90')
        self.assertEqual(str(contrato.contract_start), '2026-01-01')
        self.assertEqual(str(contrato.contract_end), '2026-12-31')
        self.assertEqual(contrato.payment_method, 'Cartao')
        self.assertEqual(contrato.card_final, '1234')
        self.assertEqual(contrato.payment_schedule, ContractEntry.PaymentSchedule.MENSAL)
        self.assertEqual(contrato.created_by, self.ti_user)

    def test_ti_can_create_contrato_with_multiple_attachments(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_contratos'),
            data={
                'name': 'Contrato com anexos',
                'notes': 'Contrato com documentos complementares.',
                'amount': '850.00',
                'contract_start': '2026-01-01',
                'contract_end': '2026-12-31',
                'payment_method': 'Boleto',
                'card_final': '',
                'payment_schedule': ContractEntry.PaymentSchedule.ANUAL,
                'attachments': [
                    ContentFile(b'contrato', name='contrato.pdf'),
                    ContentFile(b'aditivo', name='aditivo.pdf'),
                ],
            },
        )

        self.assertRedirects(response, reverse('chamados_contratos'))
        contrato = ContractEntry.objects.get(name='Contrato com anexos')
        self.assertEqual(contrato.attachments.count(), 2)

    def test_ti_can_create_contrato_anual(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_contratos'),
            data={
                'name': 'Contrato antivirus',
                'notes': 'Renovacao anual.',
                'amount': '1200.00',
                'contract_start': '2026-01-01',
                'contract_end': '2026-12-31',
                'payment_method': 'Boleto',
                'card_final': '',
                'payment_schedule': 'anual',
            },
        )

        self.assertRedirects(response, reverse('chamados_contratos'))
        contrato = ContractEntry.objects.get(name='Contrato antivirus')
        self.assertEqual(contrato.payment_schedule, ContractEntry.PaymentSchedule.ANUAL)

    def test_ti_can_create_contrato_with_brazilian_amount_mask(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_contratos'),
            data={
                'name': 'Contrato com mascara',
                'notes': 'Valor digitado no formato brasileiro.',
                'amount': '2.499,90',
                'contract_start': '2026-01-01',
                'contract_end': '2026-12-31',
                'payment_method': 'Boleto',
                'card_final': '',
                'payment_schedule': 'mensal',
            },
        )

        self.assertRedirects(response, reverse('chamados_contratos'))
        contrato = ContractEntry.objects.get(name='Contrato com mascara')
        self.assertEqual(str(contrato.amount), '2499.90')

    def test_contratos_page_displays_amount_in_brazilian_format(self):
        ContractEntry.objects.create(
            name='Contrato exibicao',
            notes='',
            amount='2499.90',
            contract_start=date(2026, 1, 1),
            contract_end=date(2026, 12, 31),
            payment_method='Boleto',
            payment_schedule=ContractEntry.PaymentSchedule.MENSAL,
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_contratos'))

        self.assertContains(response, 'R$ 2.499,90')

    def test_ti_can_attach_multiple_files_to_existing_contract(self):
        contrato = ContractEntry.objects.create(
            name='Contrato sem anexo',
            notes='',
            amount='350.00',
            contract_start=date(2026, 1, 1),
            contract_end=date(2026, 12, 31),
            payment_method='Boleto',
            payment_schedule=ContractEntry.PaymentSchedule.MENSAL,
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_contratos_attachment', args=[contrato.id]),
            data={
                'attachments': [
                    ContentFile(b'contrato-anexo', name='contrato.pdf'),
                    ContentFile(b'aditivo-anexo', name='aditivo.pdf'),
                ],
            },
        )

        self.assertRedirects(response, reverse('chamados_contratos'))
        self.assertEqual(contrato.attachments.count(), 2)

        response = self.client.get(reverse('chamados_contratos'))
        self.assertContains(response, 'Abrir 1')
        self.assertContains(response, 'Abrir 2')

    def test_ti_can_edit_existing_contract_data(self):
        contrato = ContractEntry.objects.create(
            name='Contrato antigo',
            notes='Dados antigos.',
            amount='350.00',
            contract_start=date(2026, 1, 1),
            contract_end=date(2026, 12, 31),
            payment_method='Boleto',
            payment_schedule=ContractEntry.PaymentSchedule.MENSAL,
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_contratos'),
            data={
                'mode': 'update_contract',
                'contract_id': contrato.id,
                'name': 'Contrato atualizado',
                'notes': 'Dados novos.',
                'amount': '1.200,50',
                'contract_start': '2026-04-01',
                'contract_end': '2027-03-31',
                'payment_method': 'Pix',
                'card_final': '',
                'payment_schedule': ContractEntry.PaymentSchedule.ANUAL,
            },
        )

        self.assertRedirects(response, reverse('chamados_contratos'))
        contrato.refresh_from_db()
        self.assertEqual(contrato.name, 'Contrato atualizado')
        self.assertEqual(contrato.notes, 'Dados novos.')
        self.assertEqual(str(contrato.amount), '1200.50')
        self.assertEqual(contrato.contract_start, date(2026, 4, 1))
        self.assertEqual(contrato.contract_end, date(2027, 3, 31))
        self.assertEqual(contrato.payment_method, 'Pix')
        self.assertEqual(contrato.payment_schedule, ContractEntry.PaymentSchedule.ANUAL)

    def test_contract_duration_label_is_derived_from_dates(self):
        contrato = ContractEntry.objects.create(
            name='Contrato teste',
            notes='',
            amount='100.00',
            contract_start=date(2026, 1, 1),
            contract_end=date(2027, 1, 1),
            payment_method='Boleto',
            payment_schedule=ContractEntry.PaymentSchedule.PAGAMENTO_UNICO,
            created_by=self.ti_user,
        )

        self.assertEqual(contrato.contract_duration_label, '1 ano')

    def test_only_ti_can_access_futura_digital_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_futura_digital'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_futura_digital'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Futura Digital')

    def test_ti_can_create_futura_digital_entry(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_futura_digital'),
            data={
                'name': 'Impressora RH',
                'invoice': 'FAT-2048',
                'reference_month': '2026-04',
                'copies_count': '1875',
                'paid_amount': '1.250,40',
            },
        )

        self.assertRedirects(response, reverse('chamados_futura_digital'))
        entry = FuturaDigitalEntry.objects.get()
        self.assertEqual(entry.name, 'Impressora RH')
        self.assertEqual(entry.invoice, 'FAT-2048')
        self.assertEqual(str(entry.reference_month), '2026-04-01')
        self.assertEqual(entry.copies_count, 1875)
        self.assertEqual(str(entry.paid_amount), '1250.40')
        self.assertEqual(entry.created_by, self.ti_user)

    def test_only_ti_can_access_dicas_page(self):
        self.client.login(username='usuario.comum', password='senha@123')
        response = self.client.get(reverse('chamados_dicas'))
        self.assertRedirects(response, reverse('chamados_list'))

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.get(reverse('chamados_dicas'))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Dicas')
        self.assertContains(response, 'Power Fab nao conecta')
        self.assertContains(response, 'tip-title-highlight')
        self.assertContains(response, 'tipSearchInput')
        self.assertContains(response, 'data-search=', html=False)

    def test_ti_can_create_tip_with_attachment(self):
        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_dicas'),
            data={
                'category': TipEntry.Category.GERAL,
                'title': 'Nova dica de teste',
                'content': 'Conteudo da dica.',
                'attachment': ContentFile(b'anexo-dica', name='dica_teste.txt'),
            },
        )

        self.assertRedirects(response, reverse('chamados_dicas'))
        dica = TipEntry.objects.get(title='Nova dica de teste')
        self.assertEqual(dica.created_by, self.ti_user)
        self.assertIn('dica_teste', dica.attachment.name)

    def test_whatsapp_message_uses_legacy_template_defaults(self):
        self.normal_user.first_name = 'Cassia'
        self.normal_user.last_name = 'Estevo'
        self.normal_user.save(update_fields=['first_name', 'last_name'])
        ticket = Ticket.objects.create(
            title='Chamado WhatsApp',
            description='Teste de notificacao.',
            priority=Ticket.Priority.CRITICA,
            created_by=self.normal_user,
        )

        with self.settings(
            WHATSAPP_TEMPLATE_NEW_TICKET='🚨 {urgencia} - {solicitante}\n📄 {title}'
        ):
            from chamados.whatsapp import render_new_ticket_message

            message = render_new_ticket_message(ticket)

        self.assertIn('🚨 Critica - Cassia Estevo', message)
        self.assertIn('📄 Chamado WhatsApp', message)

    def test_whatsapp_message_humanizes_username_and_newline_template(self):
        ticket = Ticket.objects.create(
            title='Teste WhatsApp',
            description='Teste de notificacao.',
            priority=Ticket.Priority.ALTA,
            created_by=self.normal_user,
        )

        with self.settings(
            WHATSAPP_TEMPLATE_NEW_TICKET='🚨 {urgencia} - {solicitante}\\n📄 {title}'
        ):
            from chamados.whatsapp import render_new_ticket_message

            message = render_new_ticket_message(ticket)

        self.assertEqual(message, '🚨 Alta - Usuario Comum\n📄 Teste WhatsApp')

    def test_whatsapp_notifications_detect_wapi_provider(self):
        with self.settings(
            WHATSAPP_NOTIFICATIONS_ENABLED=True,
            WHATSAPP_GROUP_JID='120363421981424263@g.us',
            WAPI_TOKEN='token-wapi',
            WAPI_INSTANCE='instance-01',
            WHATSAPP_WEBHOOK_URL='',
            WHATSAPP_PROVIDER='',
        ):
            from chamados.whatsapp import active_provider, notifications_enabled

            self.assertEqual(active_provider(), 'wapi')
            self.assertTrue(notifications_enabled())

    def test_whatsapp_notifications_send_via_wapi(self):
        ticket = Ticket.objects.create(
            title='Chamado WAPI',
            description='Teste de envio para W-API.',
            priority=Ticket.Priority.ALTA,
            created_by=self.normal_user,
        )

        response = MagicMock()
        response.status = 200
        response.read.return_value = json.dumps({'status': 'success', 'messageId': 'abc123'}).encode('utf-8')
        mocked_urlopen = MagicMock()
        mocked_urlopen.return_value.__enter__.return_value = response

        with self.settings(
            WHATSAPP_NOTIFICATIONS_ENABLED=True,
            WHATSAPP_GROUP_JID='120363421981424263@g.us',
            WHATSAPP_SEND_GROUP_ON_NEW_TICKET=True,
            WAPI_TOKEN='token-wapi',
            WAPI_INSTANCE='instance-01',
            WAPI_BASE_URL='https://api.w-api.app/v1',
            WHATSAPP_PROVIDER='wapi',
            WHATSAPP_WEBHOOK_URL='',
        ), patch('chamados.whatsapp.request.urlopen', mocked_urlopen):
            from chamados.whatsapp import notify_group_new_ticket

            sent = notify_group_new_ticket(ticket)

        self.assertTrue(sent)
        req = mocked_urlopen.call_args.args[0]
        self.assertIn('message/send-text?instanceId=instance-01', req.full_url)
        self.assertEqual(req.headers['Authorization'], 'Bearer token-wapi')
        payload = json.loads(req.data.decode('utf-8'))
        self.assertEqual(payload['token'], 'token-wapi')
        self.assertEqual(payload['phone'], '120363421981424263@g.us')
        self.assertIn('Chamado WAPI', payload['message'])

    def test_whatsapp_timeout_tuple_is_normalized_for_urllib(self):
        from chamados.whatsapp import _normalize_timeout

        self.assertEqual(_normalize_timeout((6.0, 20.0)), 20.0)
        self.assertEqual(_normalize_timeout(10), 10)

    def test_ti_can_update_tip(self):
        dica = TipEntry.objects.create(
            category=TipEntry.Category.GERAL,
            title='Dica antiga',
            content='Conteudo antigo.',
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(
            reverse('chamados_dicas_update', args=[dica.id]),
            data={
                'edit_tip-category': TipEntry.Category.RESOLUCAO,
                'edit_tip-title': 'Dica atualizada',
                'edit_tip-content': 'Conteudo atualizado.',
            },
        )

        self.assertRedirects(response, reverse('chamados_dicas'))
        dica.refresh_from_db()
        self.assertEqual(dica.category, TipEntry.Category.RESOLUCAO)
        self.assertEqual(dica.title, 'Dica atualizada')
        self.assertEqual(dica.content, 'Conteudo atualizado.')

    def test_only_fabiano_can_delete_tip(self):
        dica = TipEntry.objects.create(
            category=TipEntry.Category.GERAL,
            title='Dica para apagar',
            content='Conteudo removivel.',
            created_by=self.ti_user,
        )

        self.client.login(username='usuario.ti', password='senha@123')
        response = self.client.post(reverse('chamados_dicas_delete', args=[dica.id]), follow=True)
        self.assertContains(response, 'Somente fabiano.polone pode apagar dicas.')
        self.assertTrue(TipEntry.objects.filter(id=dica.id).exists())

        self.client.logout()
        self.client.login(username='fabiano.polone', password='senha@123')
        response = self.client.post(reverse('chamados_dicas_delete', args=[dica.id]))
        self.assertRedirects(response, reverse('chamados_dicas'))
        self.assertFalse(TipEntry.objects.filter(id=dica.id).exists())
